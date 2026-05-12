import numpy as np, pandas as pd, logging, warnings, glob, json, sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger("quant_pipeline")

# === CONFIG ===
@dataclass
class QuantConfig:
    LOOKBACK_DAYS: int = 252; MIN_HISTORY: int = 120; RISK_FREE_RATE: float = 0.045
    VOL_WINDOW_SHORT: int = 10; VOL_WINDOW_MED: int = 20; VOL_WINDOW_LONG: int = 60
    GARCH_P: int = 1; GARCH_Q: int = 1
    ARIMA_MAX_ORDER: int = 4; HMM_N_STATES: int = 3; HMM_N_ITER: int = 100
    ATR_PERIOD: int = 14; MAX_POSITION_RISK: float = 0.02; ACCOUNT_SIZE: float = 500_000_000
    FORECAST_HORIZON: int = 10; MONTE_CARLO_SIMS: int = 1000; CONFIDENCE_LEVEL: float = 0.95
    NORMALITY_ALPHA: float = 0.05
    # === T+2.5 CONSTRAINT (HOSE) ===
    SETTLEMENT_DAYS: float = 2.5
    MIN_HOLD_SESSIONS: int = 3
    SWING_HORIZON_MIN: int = 7
    SWING_HORIZON_MAX: int = 20
    SWING_DEFAULT: int = 10
    TICK_SIZE_RULES: dict = field(default_factory=lambda: {
        10000: 10, 50000: 50, 999999999: 100
    })
    # === V4 NEW CONFIG ===
    HMM_N_INITS: int = 10          # Multiple random inits for robustness
    HMM_VOL_WINDOW: int = 10       # Rolling vol window for HMM features
    HMM_MOM_WINDOW: int = 5        # Short momentum window for HMM features
    SR_EXTREMA_ORDER: int = 3      # argrelextrema order for swing detection
    # === V4.1 BUG FIXES ===
    VNI_CRASH_1D_PCT: float = -0.02     # 1-day drop to flag crash
    VNI_CRASH_5D_PCT: float = -0.05     # 5-day drop to flag crash
    CRISIS_VOL_MULT: float = 2.0        # Multiply GARCH vol in crisis
    FORECAST_CAP_CRISIS: float = 2.0    # Max forecast % in crisis regime
    ABSOLUTE_SCORE_ANCHOR: float = 0.5  # Blend: 50% cross-sectional + 50% absolute
    MC_SEED_PER_SYMBOL: bool = True     # Deterministic per-symbol seed

CFG = QuantConfig()

# ==============================================================
# BRIDGE: Doc Excel tu anhson.py + Lay OHLCV data
# V4.2 — Multi-source fallback: KBS (primary) → MSN (backup)
# VCI & TCBS deprecated — removed
# ==============================================================

class ScreenerBridge:
    """Cau noi giua anhson.py (Screener) va Quant Pipeline.

    V4.2: Multi-source fallback chain.
      - Primary: KBS  (KB Securities — ổn định cho HOSE daily)
      - Fallback: MSN (Microsoft — dự phòng khi KBS fail)
      - VCI và TCBS: đã bị remove (không còn hỗ trợ)

    Design:
      - Dùng `from vnstock import Quote` (API mới) thay vì
        `Vnstock().stock(...).quote` (API cũ đã deprecated).
      - Với mỗi symbol: thử SOURCE_CHAIN[0] trước; nếu data không
        hợp lệ (empty / < MIN_BARS phiên / exception), fallback tiếp theo.
      - Schema sau normalize: index=DatetimeIndex, cols=[open,high,low,close,volume].
    """

    # Thứ tự ưu tiên — swap phần tử để đổi primary/fallback
    SOURCE_CHAIN = ['KBS', 'MSN']
    MIN_BARS = 30         # Số phiên tối thiểu để coi là hợp lệ
    RETRY_SLEEP = 0.8     # Sleep giữa các lần fallback (giây)

    def __init__(self, source=None):
        """
        Args:
            source: None  -> dùng full fallback chain (khuyến nghị).
                    'KBS' -> ép chỉ dùng KBS, không fallback.
                    'MSN' -> ép chỉ dùng MSN, không fallback.
        """
        if source is None:
            self.chain = list(self.SOURCE_CHAIN)
        else:
            self.chain = [source.upper()]
        self._Quote = None

    @property
    def Quote(self):
        """Lazy import — tránh fail ngay khi import module nếu vnstock chưa cài."""
        if self._Quote is None:
            from vnstock import Quote
            self._Quote = Quote
        return self._Quote

    # ---------- Excel screener I/O (không đổi) ----------
    def read_screener_excel(self, filepath=None):
        if filepath is None:
            filepath = self._find_latest()
        if filepath is None:
            log.error("Khong tim thay file Excel tu screener!")
            return [], pd.DataFrame()
        log.info(f"Doc screener Excel: {filepath}")
        try:
            df = pd.read_excel(filepath)
            sym_col = None
            for col in df.columns:
                if any(kw in str(col).lower() for kw in ['ma', 'symbol', 'ticker', 'ma ck']):
                    sym_col = col; break
            if sym_col is None: sym_col = df.columns[0]
            symbols = [s for s in df[sym_col].astype(str).str.strip().str.upper().tolist()
                       if s and len(s) == 3 and s.isalpha()]
            log.info(f"Doc duoc {len(symbols)} ma: {symbols[:10]}...")
            return symbols, df
        except Exception as e:
            log.error(f"Loi doc Excel: {e}"); return [], pd.DataFrame()

    def _find_latest(self):
        patterns = ['Bo_Loc_Co_Phieu*.xlsx', 'bo_loc*.xlsx', '*screener*.xlsx', '*Bo_Loc*.xlsx']
        files = []
        for p in patterns:
            files.extend(glob.glob(p)); files.extend(glob.glob(f'**/{p}', recursive=True))
        return max(files, key=lambda f: Path(f).stat().st_mtime) if files else None

    # ---------- OHLCV fetching — core logic mới ----------
    def _normalize_ohlcv(self, df):
        """Chuẩn hóa schema output về [open, high, low, close, volume] với DatetimeIndex.

        Xử lý mọi biến thể column naming giữa KBS / MSN.
        Trả về None nếu schema không thể normalize.
        """
        if df is None or df.empty:
            return None

        # Map column names (case-insensitive, handle cả 'time'/'date'/'trading_date')
        cm = {}
        for c in df.columns:
            cl = str(c).lower().strip()
            if 'time' in cl or 'date' in cl:
                cm[c] = 'time'
            elif cl in ('open', 'o'):
                cm[c] = 'open'
            elif cl in ('high', 'h'):
                cm[c] = 'high'
            elif cl in ('low', 'l'):
                cm[c] = 'low'
            elif cl in ('close', 'c'):
                cm[c] = 'close'
            elif 'volume' in cl or cl == 'v':
                cm[c] = 'volume'
        df = df.rename(columns=cm)

        required = ['open', 'high', 'low', 'close', 'volume']
        if not all(c in df.columns for c in required):
            return None

        # Set DatetimeIndex
        if 'time' in df.columns:
            df['time'] = pd.to_datetime(df['time'])
            df = df.set_index('time')
        elif not isinstance(df.index, pd.DatetimeIndex):
            df.index = pd.to_datetime(df.index)

        df = df.sort_index()

        # Numeric coercion + drop bad rows
        for c in required:
            df[c] = pd.to_numeric(df[c], errors='coerce')
        df = df.dropna(subset=['close'])

        return df[required] if not df.empty else None

    def _fetch_single(self, symbol, start, end, source):
        """Fetch 1 symbol từ 1 source cụ thể. Raise exception nếu fail."""
        # API mới: Quote(symbol=..., source=...).history(...)
        q = self.Quote(symbol=symbol, source=source)
        df = q.history(start=start, end=end, interval='1D')
        return self._normalize_ohlcv(df)

    def _fetch_with_fallback(self, symbol, start, end):
        """Thử lần lượt từng source trong chain. Trả về (df, source_used) hoặc (None, None)."""
        import time
        last_err = None
        for src in self.chain:
            try:
                df = self._fetch_single(symbol, start, end, src)
                if df is not None and len(df) >= self.MIN_BARS:
                    return df, src
                # Data không đủ — thử source tiếp theo
                last_err = f"insufficient bars ({0 if df is None else len(df)} < {self.MIN_BARS})"
            except Exception as e:
                last_err = str(e)
            time.sleep(self.RETRY_SLEEP)  # backoff giữa các source
        log.warning(f"  {symbol}: all sources failed — last: {last_err}")
        return None, None

    def fetch_ohlcv(self, symbols, days=252, delay=0.3):
        """Batch fetch với fallback chain cho từng symbol."""
        import time
        data = {}
        end = datetime.now().strftime('%Y-%m-%d')
        start = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        src_stats = {s: 0 for s in self.chain}  # Track source nào đã dùng

        for i, sym in enumerate(symbols):
            log.info(f"[{i+1}/{len(symbols)}] Fetching {sym}...")
            df, src_used = self._fetch_with_fallback(sym, start, end)
            if df is not None:
                data[sym] = df
                src_stats[src_used] = src_stats.get(src_used, 0) + 1
                log.info(f"  OK {sym}: {len(df)} phien [src={src_used}]")
            time.sleep(delay)

        log.info(f"Data: {len(data)}/{len(symbols)} symbols loaded")
        log.info(f"Source breakdown: {src_stats}")
        return data

    def fetch_index(self, idx='VNINDEX', days=252):
        """Fetch index với cùng fallback chain.

        Lưu ý: MSN dùng symbol format khác cho index. Nếu KBS fail và MSN không
        hiểu 'VNINDEX', pipeline vẫn chạy được vì cfg cho phép idx_df=None.
        """
        end = datetime.now().strftime('%Y-%m-%d')
        start = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
        df, src_used = self._fetch_with_fallback(idx, start, end)
        if df is not None:
            log.info(f"Index {idx}: {len(df)} phien [src={src_used}]")
            return df
        log.warning(f"Index {idx}: fetch failed on all sources")
        return None

# ==============================================================
# M1: DISTRIBUTION ANALYSIS
# ==============================================================

class DistributionAnalyzer:
    @staticmethod
    def full_test(prices):
        from scipy import stats
        ret = prices.pct_change().dropna().values
        if len(ret) < 30: return {'error': 'Need 30+ obs'}
        r = {'n': len(ret), 'mean': round(float(np.mean(ret)),6), 'std': round(float(np.std(ret)),6),
             'skewness': round(float(stats.skew(ret)),4), 'excess_kurtosis': round(float(stats.kurtosis(ret)),4)}
        jb_s, jb_p = stats.jarque_bera(ret)
        sw_s, sw_p = stats.shapiro(ret[:5000])
        ad = stats.anderson(ret, dist='norm')
        dp_s, dp_p = stats.normaltest(ret) if len(ret) >= 20 else (0, 1)
        mu, sig = np.mean(ret), np.std(ret)
        ks_s, ks_p = stats.kstest(ret, 'norm', args=(mu, sig))
        tests = {
            'jarque_bera': {'stat': round(float(jb_s),4), 'p': round(float(jb_p),6), 'reject': bool(jb_p < CFG.NORMALITY_ALPHA)},
            'shapiro_wilk': {'stat': round(float(sw_s),6), 'p': round(float(sw_p),6), 'reject': bool(sw_p < CFG.NORMALITY_ALPHA)},
            'anderson_darling': {'stat': round(float(ad.statistic),4), 'crit_5pct': round(float(ad.critical_values[2]),4),
                                 'reject': bool(ad.statistic > ad.critical_values[2])},
            'dagostino': {'stat': round(float(dp_s),4), 'p': round(float(dp_p),6), 'reject': bool(dp_p < CFG.NORMALITY_ALPHA)},
            'ks_test': {'stat': round(float(ks_s),4), 'p': round(float(ks_p),6), 'reject': bool(ks_p < CFG.NORMALITY_ALPHA)},
        }
        r['tests'] = tests
        ek = stats.kurtosis(ret)
        p2 = np.mean(np.abs(ret) > 2*sig); n2 = 2*(1-stats.norm.cdf(2))
        p3 = np.mean(np.abs(ret) > 3*sig); n3 = 2*(1-stats.norm.cdf(3))
        try:
            ar = np.abs(ret); th = np.quantile(ar, 0.95); exc = ar[ar > th]
            hill = len(exc)/np.sum(np.log(exc/th)) if len(exc)>=5 and np.sum(np.log(exc/th))>0 else None
        except: hill = None
        r['fat_tail'] = {
            'excess_kurtosis': round(float(ek),4),
            'severity': 'EXTREME' if ek>10 else 'STRONG' if ek>5 else 'MODERATE' if ek>1 else 'MILD' if ek>0 else 'THIN',
            'tail_ratio_2sigma': round(float(p2/n2),2) if n2>0 else 1,
            'tail_ratio_3sigma': round(float(p3/n3),2) if n3>0 else 1,
            'prob_beyond_3sigma_pct': round(float(p3)*100,3),
            'hill_index': round(float(hill),3) if hill else None,
        }
        norm_p = stats.norm.fit(ret); norm_aic = 4 - 2*np.sum(stats.norm.logpdf(ret, *norm_p))
        t_p = stats.t.fit(ret); t_aic = 6 - 2*np.sum(stats.t.logpdf(ret, *t_p))
        lap_p = stats.laplace.fit(ret); lap_aic = 4 - 2*np.sum(stats.laplace.logpdf(ret, *lap_p))
        fits = {'normal': norm_aic, 'student_t': t_aic, 'laplace': lap_aic}
        best = min(fits, key=fits.get)
        r['best_fit'] = {'winner': best, 'aic': {k: round(v,2) for k,v in fits.items()}, 't_df': round(float(t_p[0]),2)}
        reject_n = sum(t['reject'] for t in tests.values())
        r['verdict'] = {
            'reject_count': reject_n, 'is_gaussian': reject_n <= 1,
            'conclusion': ('Returns GAN Gaussian' if reject_n <= 1 else
                          'Returns KHONG Gaussian - can models robust' if reject_n <= 3 else
                          'Returns RAT KHAC Gaussian - fat-tail manh'),
            'model_rec': ('Standard OK' if reject_n <= 1 else
                         'Student-t VaR, GARCH-t, Historical Sim' if reject_n <= 3 else
                         'EVT, GPD tail, non-parametric'),
        }
        return r

# ==============================================================
# M2: RETURN STATISTICS
# ==============================================================

class StatEngine:
    @staticmethod
    def returns(prices):
        if len(prices)<20: return {}
        ret = prices.pct_change().dropna()
        mr, sr = ret.mean(), ret.std()
        ar, av = mr*252, sr*np.sqrt(252)
        conf = CFG.CONFIDENCE_LEVEL
        var_p = np.percentile(ret, (1-conf)*100)
        cvar_p = ret[ret<=var_p].mean() if len(ret[ret<=var_p])>0 else var_p
        up, dn = ret[ret>0], ret[ret<0]
        wr = len(up)/len(ret)
        aw = up.mean() if len(up)>0 else 0
        al = abs(dn.mean()) if len(dn)>0 else 1e-9
        pf = aw/al
        ex = ar - CFG.RISK_FREE_RATE
        sharpe = ex/av if av>0 else 0
        dv = dn.std()*np.sqrt(252) if len(dn)>1 else av
        sortino = ex/dv if dv>0 else 0
        cum = (1+ret).cumprod(); dd = (cum - cum.cummax())/cum.cummax()
        mdd = dd.min()
        calmar = ar/abs(mdd) if mdd!=0 else 0
        return {'mean_daily_pct': round(mr*100,4), 'ann_return_pct': round(ar*100,2),
                'ann_vol_pct': round(av*100,2), 'skewness': round(ret.skew(),3),
                'kurtosis': round(ret.kurtosis(),3), 'is_fat_tail': bool(ret.kurtosis()>3),
                f'VaR_{int(conf*100)}': round(var_p*100,3),
                f'CVaR_{int(conf*100)}': round(cvar_p*100,3) if not pd.isna(cvar_p) else None,
                'max_dd_pct': round(mdd*100,2), 'sharpe': round(sharpe,3),
                'sortino': round(sortino,3), 'calmar': round(calmar,3),
                'win_rate_pct': round(wr*100,1), 'profit_factor': round(pf,3),
                'avg_win_pct': round(aw*100,3), 'avg_loss_pct': round(al*100,3), 'n': len(ret)}

    @staticmethod
    def vol_regime(prices):
        ret = prices.pct_change().dropna()
        v10 = ret.rolling(CFG.VOL_WINDOW_SHORT).std()*np.sqrt(252)
        v20 = ret.rolling(CFG.VOL_WINDOW_MED).std()*np.sqrt(252)
        v60 = ret.rolling(CFG.VOL_WINDOW_LONG).std()*np.sqrt(252)
        c10,c20,c60 = v10.iloc[-1],v20.iloc[-1],v60.iloc[-1]
        vr = c10/c60 if c60>0 else 1
        vp = (v20<c20).mean()*100 if len(v20)>=60 else 50
        rg = 'CONTRACTION' if vr<0.7 else ('EXPANSION' if vr>1.3 else 'NORMAL')
        return {'vol_10d': round(c10*100,2) if not pd.isna(c10) else None,
                'vol_20d': round(c20*100,2) if not pd.isna(c20) else None,
                'vol_60d': round(c60*100,2) if not pd.isna(c60) else None,
                'vol_ratio': round(vr,3), 'vol_pctile': round(vp,1), 'regime': rg}

    @staticmethod
    def autocorr(prices, max_lag=10):
        ret = prices.pct_change().dropna()
        if len(ret)<max_lag+10: return {'behavior': 'INSUFFICIENT'}
        ac = {f'lag_{i}': round(ret.autocorr(lag=i),4) for i in range(1, max_lag+1)}
        avg = np.mean([ac.get(f'lag_{i}',0) for i in range(1,4)])
        th = 2/np.sqrt(len(ret))
        beh = 'MOMENTUM' if avg>th else ('MEAN_REVERSION' if avg<-th else 'RANDOM_WALK')
        return {'ac': ac, 'avg_short': round(avg,4), 'threshold': round(th,4), 'behavior': beh}

# ==============================================================
# M3: ARIMA
# ==============================================================

class ARIMAEngine:
    @staticmethod
    def fit(prices, max_p=4, max_q=4):
        try:
            from statsmodels.tsa.arima.model import ARIMA
            from statsmodels.tsa.stattools import adfuller
        except ImportError: return {'error': 'pip install statsmodels'}
        ret = prices.pct_change().dropna()
        if len(ret)<60: return {'error': 'Need 60+ obs'}
        adf_s, adf_p = adfuller(ret.values, maxlag=20)[:2]
        is_stat = adf_p < 0.05
        r = {'stationarity': {'adf_stat': round(adf_s,4), 'adf_p': round(adf_p,6), 'stationary': is_stat}}
        best_aic, best_order, best_m = np.inf, (0,0,0), None
        for p in range(0, max_p+1):
            for q in range(0, max_q+1):
                if p==0 and q==0: continue
                try:
                    m = ARIMA(ret.values, order=(p,0,q)).fit()
                    if m.aic < best_aic: best_aic, best_order, best_m = m.aic, (p,0,q), m
                except: pass
        if best_m is None: r['error'] = 'No ARIMA fit'; return r
        r['best'] = {'order': f'ARIMA{best_order}', 'aic': round(best_aic,2), 'bic': round(best_m.bic,2)}
        coefs = {}
        for nm, val, pv in zip(best_m.param_names, best_m.params, best_m.pvalues):
            coefs[nm] = {'val': round(float(val),6), 'p': round(float(pv),4), 'sig': bool(pv<0.05)}
        r['coefficients'] = coefs
        fc = best_m.forecast(steps=CFG.FORECAST_HORIZON)
        lp = prices.iloc[-1]; fp = [lp]
        for rt in fc: fp.append(fp[-1]*(1+rt))
        r['forecast'] = {'prices': [round(p,0) for p in fp[1:]], 'returns_pct': [round(x*100,3) for x in fc]}
        return r

# ==============================================================
# M4: GARCH / EGARCH
# ==============================================================

class GARCHEngine:
    @staticmethod
    def fit(prices, p=1, q=1):
        try:
            from arch import arch_model
        except ImportError: return {'error': 'pip install arch'}
        ret = prices.pct_change().dropna()*100
        if len(ret)<100: return {'error': 'Need 100+ obs'}
        r = {}
        try:
            m = arch_model(ret, vol='Garch', p=p, q=q, dist='t').fit(disp='off')
            alpha = float(m.params.get('alpha[1]',0)); beta = float(m.params.get('beta[1]',0))
            persist = alpha+beta
            fc = m.forecast(horizon=CFG.FORECAST_HORIZON)
            cvol = np.sqrt(fc.variance.iloc[-1].values)
            r['garch'] = {
                'model': f'GARCH({p},{q})-t', 'aic': round(float(m.aic),2),
                'alpha': round(alpha,4), 'beta': round(beta,4), 'persistence': round(persist,4),
                'half_life': round(np.log(0.5)/np.log(persist),1) if 0<persist<1 else None,
                'persist_label': ('IGARCH-like' if persist>0.97 else 'High' if persist>0.9 else 'Moderate' if persist>0.7 else 'Low'),
                'current_vol_pct': round(float(np.sqrt(m.conditional_volatility.iloc[-1])),3),
                'forecast_vol': [round(float(v),3) for v in cvol],
                'shock_sensitivity': 'HIGH' if alpha>0.15 else ('MODERATE' if alpha>0.08 else 'LOW'),
            }
            params = {}
            for nm in m.params.index:
                params[nm] = {'val': round(float(m.params[nm]),6), 'p': round(float(m.pvalues[nm]),4)}
            r['garch']['params'] = params
        except Exception as e: r['garch'] = {'error': str(e)}
        try:
            em = arch_model(ret, vol='EGARCH', p=1, q=1, dist='t').fit(disp='off')
            gamma = float(em.params.get('gamma[1]',0))
            r['egarch'] = {'aic': round(float(em.aic),2), 'gamma': round(gamma,4),
                           'leverage': bool(gamma<0),
                           'note': 'Tin xau tang vol MANH hon tin tot' if gamma<-0.05 else 'Leverage effect yeu'}
        except Exception as e: r['egarch'] = {'error': str(e)}
        return r

# ==============================================================
# M5: HMM REGIME — V4: MULTIVARIATE + MULTI-INIT + BIC SELECT
# ==============================================================
#
# FIX #2: Single-feature HMM too noisy for position trading.
#   -> 4 features: [return, rolling_vol, volume_ratio, short_momentum]
#   -> N random inits, select best model by BIC
#   -> Robust state labeling via return-dimension ranking
# ==============================================================

class HMMEngine:
    @staticmethod
    def _build_features(df):
        """Build multivariate feature matrix for HMM.
        4 features standardized to prevent scale dominance:
          1. Daily return         — directional
          2. Rolling vol (10d)    — vol regime
          3. Volume ratio (5/20)  — participation
          4. Short momentum (5d)  — trend smoothing
        """
        close = df['close']; volume = df['volume']
        ret = close.pct_change()
        rolling_vol = ret.rolling(CFG.HMM_VOL_WINDOW).std()
        vol_ma5 = volume.rolling(5).mean()
        vol_ma20 = volume.rolling(20).mean()
        vol_ratio = vol_ma5 / vol_ma20.replace(0, np.nan)
        short_mom = ret.rolling(CFG.HMM_MOM_WINDOW).mean()
        feat_df = pd.DataFrame({
            'return': ret, 'rolling_vol': rolling_vol,
            'vol_ratio': vol_ratio, 'short_mom': short_mom,
        }).dropna()
        if len(feat_df) < 100: return None, None
        values = feat_df.values.copy()
        means = values.mean(axis=0); stds = values.std(axis=0)
        stds[stds == 0] = 1.0
        values_z = (values - means) / stds
        return values_z, feat_df.index

    @staticmethod
    def fit(df, n=3, symbol=''):
        try:
            from hmmlearn.hmm import GaussianHMM
        except ImportError: return HMMEngine._fallback(df)
        features, feat_idx = HMMEngine._build_features(df)
        if features is None: return HMMEngine._fallback(df)
        try:
            best_model, best_bic = None, np.inf
            n_samples, n_features = features.shape
            # V4.1: Symbol-based seed for stability across runs
            sym_seed = abs(hash(symbol)) % (2**31) if symbol else 0
            for init_i in range(CFG.HMM_N_INITS):
                try:
                    m = GaussianHMM(n_components=n, covariance_type='full',
                                    n_iter=CFG.HMM_N_ITER, random_state=sym_seed + init_i, tol=1e-4)
                    m.fit(features)
                    ll = m.score(features)
                    n_params = (n*(n-1) + n*n_features + n*n_features*(n_features+1)//2 + (n-1))
                    bic = -2*ll + n_params * np.log(n_samples)
                    if bic < best_bic: best_bic, best_model = bic, m
                except: continue
            if best_model is None: return HMMEngine._fallback(df)
            m = best_model; states = m.predict(features)
            means = m.means_; ret_means = means[:, 0]
            order = np.argsort(ret_means)
            names = {order[0]: 'BEAR', order[1]: 'SIDEWAY', order[2]: 'BULL'}
            emj = {'BEAR': '🔴', 'SIDEWAY': '🟡', 'BULL': '🟢'}
            info = {}
            for i in range(n):
                nm = names.get(i, '?')
                state_mask = states == i
                state_ret = df['close'].pct_change().reindex(feat_idx).values[state_mask]
                state_ret = state_ret[~np.isnan(state_ret)]
                info[nm] = {
                    'mean_daily_pct': round(float(np.mean(state_ret))*100, 4) if len(state_ret)>0 else 0,
                    'ann_ret_pct': round(float(np.mean(state_ret))*252*100, 2) if len(state_ret)>0 else 0,
                    'ann_vol_pct': round(float(np.std(state_ret))*np.sqrt(252)*100, 2) if len(state_ret)>1 else 0,
                    'pct_time': round(float(np.mean(state_mask))*100, 1),
                    'avg_vol_regime': round(float(means[i, 1]), 2),
                    'avg_volume_activity': round(float(means[i, 2]), 2),
                }
            cs = states[-1]; cn = names.get(cs, '?')
            probs = m.predict_proba(features)[-1]
            np_d = {names.get(i,'?'): round(float(probs[i])*100,1) for i in range(n)}
            trans = {}
            for i in range(n):
                f = names.get(i,'?')
                trans[f] = {names.get(j,'?'): round(float(m.transmat_[i,j])*100,1) for j in range(n)}
            return {
                'current': f"{emj.get(cn,'')} {cn}", 'prob_pct': round(float(max(probs))*100,1),
                'state_probs': np_d, 'states': info, 'transitions': trans,
                'recent_20': [names.get(s,'?') for s in states[-20:]],
                'warning': '⚠️ Regime co the doi' if max(probs)<0.6 else 'On dinh',
                'method': 'multivariate_hmm', 'n_features': features.shape[1],
                'n_inits': CFG.HMM_N_INITS, 'bic': round(best_bic, 2),
            }
        except Exception as e:
            log.warning(f"HMM fit error: {e}"); return HMMEngine._fallback(df)

    @staticmethod
    def _fallback(df):
        prices = df['close'] if isinstance(df, pd.DataFrame) else df
        ma20, ma50 = prices.rolling(20).mean(), prices.rolling(50).mean()
        p, m2, m5 = prices.iloc[-1], ma20.iloc[-1], ma50.iloc[-1]
        r20 = prices.iloc[-1]/prices.iloc[-20]-1 if len(prices)>=20 else 0
        st = '🟢 BULL' if p>m2>m5 and r20>0.02 else ('🔴 BEAR' if p<m2<m5 and r20<-0.02 else '🟡 SIDEWAY')
        return {'method': 'fallback', 'current': st, 'note': 'pip install hmmlearn cho HMM'}

# ==============================================================
# M6: ALPHA SIGNALS
# ==============================================================

class AlphaEngine:
    @staticmethod
    def extract(df, index_df=None):
        if len(df)<30: return {}
        c = df['close']; v = df['volume']
        sig = {}
        delta = c.diff(); gain = delta.where(delta>0,0).rolling(14).mean()
        loss = -delta.where(delta<0,0).rolling(14).mean()
        rsi = 100 - 100/(1+gain/loss.replace(0,1e-9)); crsi = rsi.iloc[-1]
        ema12, ema26 = c.ewm(span=12).mean(), c.ewm(span=26).mean()
        macd_h = (ema12-ema26) - (ema12-ema26).ewm(span=9).mean()
        roc10 = (c.iloc[-1]/c.iloc[-10]-1)*100 if len(c)>=10 else 0
        ms = 0
        if crsi>70: ms-=0.3
        elif crsi<30: ms+=0.3
        elif crsi>50: ms+=0.1
        else: ms-=0.1
        if roc10>5: ms+=0.3
        elif roc10<-5: ms-=0.3
        if macd_h.iloc[-1]>0: ms+=0.2
        else: ms-=0.2
        sig['momentum'] = {'rsi': round(float(crsi),1), 'roc_10d': round(roc10,2),
                          'macd_hist': round(float(macd_h.iloc[-1]),2), 'score': round(np.clip(ms,-1,1),3)}
        ma20 = c.rolling(20).mean(); std20 = c.rolling(20).std()
        z = (c.iloc[-1]-ma20.iloc[-1])/std20.iloc[-1] if std20.iloc[-1]>0 else 0
        bb_u = ma20.iloc[-1]+2*std20.iloc[-1]; bb_l = ma20.iloc[-1]-2*std20.iloc[-1]
        mrs = 0
        if z>2: mrs-=0.5
        elif z<-2: mrs+=0.5
        elif abs(z)>1: mrs-=0.2*np.sign(z)
        sig['mean_reversion'] = {'z_score': round(float(z),3), 'bb_upper': round(float(bb_u),0),
                                 'bb_lower': round(float(bb_l),0), 'score': round(np.clip(mrs,-1,1),3)}
        obv = (np.sign(c.diff())*v).cumsum()
        obv_up = obv.iloc[-1]>obv.iloc[-20] if len(obv)>=20 else True
        vr = v.rolling(5).mean().iloc[-1] / v.rolling(20).mean().iloc[-1] if v.rolling(20).mean().iloc[-1]>0 else 1
        vs = (0.3 if obv_up else -0.3) + (0.2 if vr>1.5 else (-0.2 if vr<0.5 else 0))
        sig['volume'] = {'obv_trend': 'UP' if obv_up else 'DOWN', 'vol_ratio_5_20': round(float(vr),2),
                        'score': round(np.clip(vs,-1,1),3)}
        if index_df is not None and 'close' in index_df.columns:
            # Deduplicate index BEFORE reindex to avoid ValueError
            index_df = index_df[~index_df.index.duplicated(keep='last')]
            c = c[~c.index.duplicated(keep='last')]
            ic = index_df['close'].reindex(c.index).ffill()
            if len(c)>=20 and len(ic)>=20:
                sr20 = c.pct_change(20).iloc[-1] - ic.pct_change(20).iloc[-1]
                sr, ir = c.pct_change().dropna(), ic.pct_change().reindex(c.pct_change().dropna().index).dropna()
                common = sr.index.intersection(ir.index)
                if len(common)>=20:
                    cov = np.cov(sr[common].values, ir[common].values)
                    beta = cov[0,1]/cov[1,1] if cov[1,1]>0 else 1
                    alpha_ann = (sr[common].mean()-beta*ir[common].mean())*252
                else: beta, alpha_ann = 1, 0
                sig['cross_sectional'] = {'rs_20d_pct': round(float(sr20)*100,2), 'beta': round(float(beta),3),
                                          'alpha_ann_pct': round(float(alpha_ann)*100,2)}
        # Composite — weights = 0.4 + 0.3 + 0.3 = 1.0
        cs = (0.4*sig.get('momentum',{}).get('score',0) +
              0.3*sig.get('mean_reversion',{}).get('score',0) +
              0.3*sig.get('volume',{}).get('score',0))
        sig['composite'] = {'alpha': round(cs,3),
                           'label': '🟢 BULLISH' if cs>0.3 else ('🔴 BEARISH' if cs<-0.3 else '⚪ NEUTRAL')}
        return sig

# ==============================================================
# M7: MARKET STRUCTURE — V4: SWING HIGH/LOW S/R
# ==============================================================
#
# FIX #4: Histogram S/R finds price frequency, not turning points.
#   -> argrelextrema for real swing high/low detection
#   -> Cluster nearby levels within ATR distance
#   -> Rank by touches + recency
# ==============================================================

class StructureEngine:
    @staticmethod
    def sr_levels(df, n=3, lb=120):
        from scipy.signal import argrelextrema
        if len(df) < 30: return {'supports': [], 'resistances': []}
        tail = df.tail(lb)
        highs = tail['high'].values; lows = tail['low'].values
        close_now = tail['close'].iloc[-1]
        order = CFG.SR_EXTREMA_ORDER
        high_idx = argrelextrema(highs, np.greater_equal, order=order)[0]
        low_idx = argrelextrema(lows, np.less_equal, order=order)[0]
        atr = RiskEng.atr(df)
        cluster_dist = max(atr * 0.5, close_now * 0.005)

        def cluster_levels(prices, indices, n_levels):
            if len(indices) == 0: return []
            n_bars = len(prices)
            levels = [(prices[idx], 1.0 + (idx / n_bars)) for idx in indices]
            levels.sort(key=lambda x: x[0])
            clusters = []; current = [levels[0]]
            for i in range(1, len(levels)):
                if levels[i][0] - current[-1][0] <= cluster_dist:
                    current.append(levels[i])
                else:
                    clusters.append(current); current = [levels[i]]
            clusters.append(current)
            result = []
            for cl in clusters:
                avg_p = np.mean([c[0] for c in cl])
                touches = len(cl)
                avg_rec = np.mean([c[1] for c in cl])
                result.append({'price': round(float(avg_p),0), 'str': round(float(touches*avg_rec),2), 'touches': touches})
            result.sort(key=lambda x: x['str'], reverse=True)
            return result[:n_levels]

        all_res = cluster_levels(highs, high_idx, n*2)
        all_sup = cluster_levels(lows, low_idx, n*2)
        supports = sorted([s for s in all_sup if s['price']<close_now], key=lambda x:x['price'], reverse=True)[:n]
        resistances = sorted([r for r in all_res if r['price']>=close_now], key=lambda x:x['price'])[:n]
        return {'price': round(close_now,0), 'supports': supports, 'resistances': resistances, 'method': 'swing_extrema'}

    @staticmethod
    def trend(df, period=20):
        if len(df)<period+1: return {}
        p = df.tail(period+1)['close'].values
        d = abs(p[-1]-p[0]); vol = np.sum(np.abs(np.diff(p)))
        er = d/vol if vol>0 else 0; slope = (p[-1]-p[0])/p[0]
        x = np.arange(len(p)); co = np.polyfit(x,p,1); tl = np.polyval(co,x)
        ss_r = np.sum((p-tl)**2); ss_t = np.sum((p-np.mean(p))**2)
        r2 = 1-ss_r/ss_t if ss_t>0 else 0
        lb = ('📈 UPTREND' if er>0.6 and slope>0 else '📉 DOWNTREND' if er>0.6 and slope<0 else
              '↗️ UP NHE' if er>0.3 and slope>0 else '↘️ DOWN NHE' if er>0.3 else '↔️ SIDEWAY')
        return {'er': round(er,4), 'r2': round(r2,4), 'slope_pct': round(slope*100,2), 'label': lb}

    @staticmethod
    def flow(df, lb=20):
        r = df.tail(lb)
        if len(r)<5: return {}
        h,l,c,v = r['high'].values, r['low'].values, r['close'].values, r['volume'].values.astype(float)
        rng = np.where((h-l)==0, 1e-9, h-l)
        mf = ((c-l)-(h-c))/rng
        cmf = np.sum(mf*v)/np.sum(v) if np.sum(v)>0 else 0
        bp = np.mean((c-l)/rng)
        lb2 = ('🟢 MUA' if cmf>0.1 and bp>0.6 else '🔴 BAN' if cmf<-0.1 else '🟡 THIEN MANH' if cmf>0 else '🟡 THIEN YEU')
        return {'cmf': round(cmf,4), 'buy_p': round(bp,3), 'label': lb2}

# ==============================================================
# M8: FORECASTING — V4: GARCH-INFORMED MONTE CARLO
# ==============================================================
#
# FIX #3: MC uses constant mu/sigma, ignoring GARCH forecast.
#   -> Accept garch_result, extract h-step forecast vols
#   -> Time-varying vol per simulation step
#   -> Fallback to historical sigma if GARCH unavailable
# ==============================================================

class FcastEngine:
    @staticmethod
    def ensemble(prices, garch_result=None, symbol='', vni_regime='NEUTRAL'):
        if len(prices)<30: return {}
        lr = np.log(prices/prices.shift(1)).dropna().values
        mu = np.mean(lr); hist_sigma = np.std(lr); lp = prices.iloc[-1]

        # V4.1 FIX #1: Per-symbol deterministic RNG instead of global seed
        sym_seed = abs(hash(symbol)) % (2**31) if symbol else 42
        rng = np.random.RandomState(sym_seed)

        H = CFG.SWING_DEFAULT; LOCK = CFG.MIN_HOLD_SESSIONS; N = CFG.MONTE_CARLO_SIMS

        # V4.1 FIX #2: Extract GARCH forecast vols with sanity bounds
        garch_vols = None
        if garch_result is not None:
            gv = garch_result.get('garch', {}).get('forecast_vol', [])
            if gv and len(gv) >= H:
                garch_vols = np.array(gv[:H]) / 100.0
            elif gv:
                extended = list(gv) + [gv[-1]] * (H - len(gv))
                garch_vols = np.array(extended[:H]) / 100.0
            if garch_vols is not None:
                # Sanity: cap vol at 3x historical, floor at 0.3x historical
                vol_floor = hist_sigma * 0.3
                vol_cap = hist_sigma * 3.0
                garch_vols = np.clip(garch_vols, vol_floor, vol_cap)
                # V4.1: Crisis vol multiplier
                if vni_regime == 'CRISIS':
                    garch_vols = garch_vols * CFG.CRISIS_VOL_MULT

        # Monte Carlo — GARCH-aware with per-symbol RNG
        sims = np.zeros((N, H))
        if garch_vols is not None:
            for step in range(H):
                z = rng.standard_normal(N)
                if step == 0:
                    sims[:, step] = lp * np.exp(mu + garch_vols[step] * z)
                else:
                    sims[:, step] = sims[:, step-1] * np.exp(mu + garch_vols[step] * z)
        else:
            sigma_use = hist_sigma
            if vni_regime == 'CRISIS': sigma_use *= CFG.CRISIS_VOL_MULT
            for i in range(N):
                sims[i] = lp * np.exp(np.cumsum(rng.normal(mu, sigma_use, H)))

        fin = sims[:,-1]
        mc = {'price': round(np.mean(fin),0), 'prob_up': round(np.mean(fin>lp)*100,1),
              'ci_lo': round(np.percentile(fin,2.5),0), 'ci_hi': round(np.percentile(fin,97.5),0),
              'vol_source': 'GARCH' if garch_vols is not None else 'historical'}

        lock_prices = sims[:, :LOCK]; lock_mins = lock_prices.min(axis=1)
        lock_dd = (lock_mins - lp) / lp * 100
        lock_risk = {
            'max_dd_lock_pct': round(np.percentile(lock_dd, 5), 2),
            'avg_dd_lock_pct': round(np.mean(lock_dd[lock_dd<0]), 2) if np.any(lock_dd<0) else 0,
            'prob_loss_in_lock_pct': round(np.mean(lock_mins < lp)*100, 1),
            'prob_loss_gt_3pct': round(np.mean(lock_dd < -3)*100, 1),
            'lock_sessions': LOCK,
        }
        unlock_prices = sims[:, LOCK-1]
        unlock = {
            'price_at_unlock': round(np.mean(unlock_prices), 0),
            'prob_up_at_unlock': round(np.mean(unlock_prices > lp)*100, 1),
            'ci_lo_unlock': round(np.percentile(unlock_prices, 10), 0),
            'ci_hi_unlock': round(np.percentile(unlock_prices, 90), 0),
        }
        ma20 = prices.rolling(20).mean(); std20 = prices.rolling(20).std()
        z = (lp-ma20.iloc[-1])/std20.iloc[-1] if std20.iloc[-1]>0 else 0
        mr_target = ma20.iloc[-1]
        mr_sessions = min(H, max(LOCK+1, int(abs(z)*3)))
        mr = {'z': round(z,3), 'target': round(mr_target,0), 'dir': 'DOWN' if z>0 else 'UP', 'est_sessions': mr_sessions}
        rd = prices.pct_change().tail(H); avg = rd.mean()
        ms = avg/rd.std() if rd.std()>0 else 0
        mom = {'strength': round(ms,3), 'proj_pct': round(avg*H*100,2),
               'signal': 'BULLISH' if ms>0.3 else ('BEARISH' if ms<-0.3 else 'NEUTRAL')}

        # V4.1 FIX #7: Balanced weights — prevent auto-switch to MR after crash
        # In crisis/bear: momentum gets MORE weight (trend-following, not mean-reverting)
        msa = abs(ms); za = abs(z)
        if vni_regime in ('CRISIS', 'BEAR'):
            # In downtrend: trust momentum > MR, MC as base
            w = (0.35, 0.15, 0.50)
        elif msa > 0.5:
            w = (0.30, 0.20, 0.50)
        elif za > 1.5:
            # Only trust MR in non-crisis environments
            w = (0.30, 0.45, 0.25)
        else:
            w = (0.50, 0.25, 0.25)

        mc_r = (np.mean(fin)/lp-1)*100; mr_r = (mr_target-lp)/lp*100; mom_r = avg*H*100
        er = w[0]*mc_r + w[1]*mr_r + w[2]*mom_r

        # V4.1: Cap forecast in crisis regime
        if vni_regime == 'CRISIS':
            er = np.clip(er, -CFG.FORECAST_CAP_CRISIS * 3, CFG.FORECAST_CAP_CRISIS)
        elif vni_regime == 'BEAR':
            er = np.clip(er, -20, CFG.FORECAST_CAP_CRISIS * 2)

        cons = '📈 TANG' if er > 0.5 else ('📉 GIAM' if er < -0.5 else '↔️ TRUNG LAP')
        dirs = []
        if mc['prob_up']>55: dirs.append('UP')
        elif mc['prob_up']<45: dirs.append('DOWN')
        if mr['dir']=='UP': dirs.append('UP')
        else: dirs.append('DOWN')
        if mom['signal']=='BULLISH': dirs.append('UP')
        elif mom['signal']=='BEARISH': dirs.append('DOWN')
        er_dir = 'UP' if er > 0 else 'DOWN'
        agree = sum(1 for d in dirs if d == er_dir)
        conf = agree/len(dirs)*100 if dirs else 33
        lock_safe = lock_risk['prob_loss_gt_3pct'] < 20
        signal_strong = abs(er) > 2 and conf >= 66

        # V4.1 FIX #3: Timing MUST respect VNI regime — circuit breaker
        if vni_regime == 'CRISIS':
            timing = '🔴 KHÔNG VÀO — CRISIS: thị trường đang trong trạng thái khẩn cấp'
        elif vni_regime == 'BEAR':
            # In BEAR: never VÀO NGAY, at most CHỜ PULLBACK
            if signal_strong and lock_safe:
                timing = '🟡 CHỜ PULLBACK — tín hiệu tốt nhưng VNI đang BEAR'
            elif signal_strong:
                timing = '🟡 CHỜ PULLBACK — tín hiệu tốt nhưng lock risk cao + VNI BEAR'
            elif lock_safe:
                timing = '🟡 THEO DÕI — lock an toàn nhưng tín hiệu yếu + VNI BEAR'
            else:
                timing = '🔴 KHÔNG VÀO — lock risk cao + tín hiệu yếu + VNI BEAR'
        else:
            # Original logic for NEUTRAL/BULL/WEAK
            if signal_strong and lock_safe:
                timing = '🟢 VÀO NGAY — lock risk thấp, tín hiệu mạnh'
            elif signal_strong and not lock_safe:
                timing = '🟡 CHỜ PULLBACK — tín hiệu tốt nhưng lock risk cao'
            elif not signal_strong and lock_safe:
                timing = '🟡 THEO DÕI — lock an toàn nhưng tín hiệu yếu'
            else:
                timing = '🔴 KHÔNG VÀO — lock risk cao + tín hiệu yếu'

        # V4.1 FIX #10: Hold plan with regime awareness
        if vni_regime in ('CRISIS', 'BEAR'):
            hold_plan = CFG.SWING_HORIZON_MIN  # Minimize exposure
        elif 'TANG' in cons and ms > 0.3:
            hold_plan = CFG.SWING_HORIZON_MAX
        elif 'TANG' in cons:
            hold_plan = CFG.SWING_DEFAULT
        else:
            hold_plan = CFG.SWING_HORIZON_MIN

        # ── MC path summary for Forecast chart export ──
        mc_median_path = np.median(sims, axis=0).tolist()
        mc_ci_upper = np.percentile(sims, 97.5, axis=0).tolist()
        mc_ci_lower = np.percentile(sims, 2.5, axis=0).tolist()

        return {
            'ensemble_ret_pct': round(er,2), 'ensemble_price': round(lp*(1+er/100),0),
            'consensus': cons, 'confidence': round(conf,0), 'horizon': H,
            'mc': mc, 'mr': mr, 'mom': mom, 'lock_risk': lock_risk, 'unlock': unlock,
            'timing': timing, 'hold_plan_sessions': hold_plan,
            'hold_plan_label': f'{hold_plan} phien (~{round(hold_plan*7/5)} ngay)',
            'mc_path': {'median': mc_median_path, 'upper': mc_ci_upper, 'lower': mc_ci_lower},
        }

# ==============================================================
# M9: RISK ENGINE
# ==============================================================

class RiskEng:
    @staticmethod
    def atr(df, per=14):
        if len(df)<per+1: return 0
        h,l,c = df['high'].values, df['low'].values, df['close'].values
        tr = np.maximum(h[1:]-l[1:], np.maximum(abs(h[1:]-c[:-1]), abs(l[1:]-c[:-1])))
        a = pd.Series(tr).rolling(per).mean().iloc[-1]
        return a if not pd.isna(a) else 0

    @staticmethod
    def tick_round(price, direction='down'):
        for threshold, tick in sorted(CFG.TICK_SIZE_RULES.items()):
            if price < threshold:
                if direction == 'down': return (price // tick) * tick
                else: return ((price // tick) + 1) * tick
        return round(price, 0)

    @staticmethod
    def stops(df):
        if len(df)<20: return {}
        e = df['close'].iloc[-1]; a = RiskEng.atr(df)
        if a <= 0: return {'entry': round(e,0), 'atr': 0, 'warning': 'ATR=0'}
        sl_swing = RiskEng.tick_round(e - 2.5*a, 'down')
        sl_wide  = RiskEng.tick_round(e - 3.5*a, 'down')
        sl_max   = RiskEng.tick_round(e - 4.5*a, 'down')
        swing_low_10 = df['low'].tail(10).min()
        swing_low_20 = df['low'].tail(20).min()
        risk = e - sl_swing
        tp_2r = RiskEng.tick_round(e + 2*risk, 'up')
        tp_3r = RiskEng.tick_round(e + 3*risk, 'up')
        tp_4r = RiskEng.tick_round(e + 4*risk, 'up')
        trail_atr = round(2*a, 0)
        if len(df) >= 10:
            lock_hist_dd = []
            closes = df['close'].values; lows = df['low'].values
            for i in range(3, len(closes)):
                entry_p = closes[i-3]
                worst_in_lock = min(lows[i-2], lows[i-1], lows[i])
                lock_hist_dd.append((worst_in_lock - entry_p) / entry_p * 100)
            mae_lock = round(np.percentile(lock_hist_dd, 10), 2) if lock_hist_dd else 0
        else: mae_lock = 0
        return {
            'entry': round(e,0), 'atr': round(a,0),
            'sl_swing': sl_swing, 'sl_wide': sl_wide, 'sl_max': sl_max,
            'swing_low_10': round(swing_low_10,0), 'swing_low_20': round(swing_low_20,0),
            'tp_2r': tp_2r, 'tp_3r': tp_3r, 'tp_4r': tp_4r, 'risk_per_share': round(risk,0),
            'trail_atr': trail_atr, 'trail_note': f'Sau T+3: trailing {trail_atr:,.0f} VND (2x ATR)',
            'mae_lock_10pct': mae_lock,
            'mae_lock_note': f'Lich su: 90% truong hop DD trong lock < {abs(mae_lock):.1f}%',
            'settlement': f'T+{CFG.SETTLEMENT_DAYS} — som nhat ban phien {CFG.MIN_HOLD_SESSIONS+1}',
        }

    @staticmethod
    def sizing(entry, stop, acct=None, hold_sessions=None):
        acct = acct or CFG.ACCOUNT_SIZE
        rps = abs(entry-stop)
        if rps<=0: return {}
        ml = acct*CFG.MAX_POSITION_RISK
        sh = (int(ml/rps)//100)*100; val = sh*entry
        pct_acct = val/acct*100
        max_pct = 15.0
        if pct_acct > max_pct:
            sh = int((acct * max_pct/100) / entry / 100) * 100
            val = sh * entry; pct_acct = val/acct*100
        hold = hold_sessions or CFG.SWING_DEFAULT
        capital_lock_days = round(hold * 7/5)
        return {'shares': sh, 'value': round(val,0), 'pct_acct': round(pct_acct,1), 'max_loss': round(ml,0),
                'capital_lock_days': capital_lock_days,
                'capital_lock_note': f'Von {val:,.0f} VND bi khoa ~{capital_lock_days} ngay'}

    @staticmethod
    def kelly(wr, aw, al):
        if al==0: return {}
        b = aw/al; p,q = wr, 1-wr; k = (p*b-q)/b
        return {'full_pct':round(max(0,k)*100,2), 'half_pct':round(max(0,k/2)*100,2),
                'edge_pct':round((p*b-q)*100,2), 'has_edge':bool(k>0)}

# ==============================================================
# ADAPTIVE SCORER — V4: CROSS-SECTIONAL Z-SCORE SCORING
# ==============================================================
#
# FIX #5: Magic numbers replaced with z-score normalization.
#   For each factor across N stocks in batch:
#     z_i = (f_i - median) / IQR   [robust z-score]
#     s_i = clip(z_i, -1, +1)      [bounded signal]
#   composite = sum(w_i * s_i), sum(w_i) = 1
#   final = 50 + composite * 40    [maps to ~[10,90]]
#   + VNI regime multiplicative adjustment
# ==============================================================

class AdaptiveScorer:
    WEIGHTS = {
        'sharpe': 0.12, 'win_rate': 0.06, 'forecast_ret': 0.14,
        'forecast_conf': 0.08, 'cmf': 0.10, 'trend_quality': 0.10,
        'alpha_composite': 0.12, 'hmm_regime': 0.08, 'lock_safety': 0.08,
        'vol_regime': 0.06, 'distribution_quality': 0.06,
    }

    @staticmethod
    def _robust_zscore(value, values_array):
        if len(values_array) < 3: return 0.0
        med = np.median(values_array)
        iqr = np.percentile(values_array, 75) - np.percentile(values_array, 25)
        if iqr < 1e-9: return 0.0
        return float(np.clip((value - med) / iqr, -1.0, 1.0))

    @staticmethod
    def extract_factors(report):
        rs = report.get('stats', {}); fc = report.get('fcast', {})
        fl = report.get('flow', {}); tr = report.get('trend', {})
        al = report.get('alpha', {}).get('composite', {}); hm = report.get('hmm', {})
        vl = report.get('vol', {}); di = report.get('dist', {}).get('verdict', {})
        lock = fc.get('lock_risk', {})
        hmm_cur = hm.get('current', '')
        hmm_num = 1.0 if 'BULL' in hmm_cur else (-1.0 if 'BEAR' in hmm_cur else 0.0)
        hmm_score = hmm_num * (hm.get('prob_pct', 50) / 100.0)
        er = tr.get('er', 0); slope = tr.get('slope_pct', 0)
        trend_q = er * np.sign(slope) if slope != 0 else 0
        lock_safety = 100 - lock.get('prob_loss_gt_3pct', 50)
        vol_score = 1.0 - vl.get('vol_ratio', 1.0)
        dist_score = 1.0 if di.get('is_gaussian', True) else -0.5
        return {
            'sharpe': rs.get('sharpe', 0),
            'win_rate': (rs.get('win_rate_pct', 50) - 50) / 50,
            'forecast_ret': fc.get('ensemble_ret_pct', 0),
            'forecast_conf': (fc.get('confidence', 50) - 50) / 50,
            'cmf': fl.get('cmf', 0), 'trend_quality': trend_q,
            'alpha_composite': al.get('alpha', 0), 'hmm_regime': hmm_score,
            'lock_safety': (lock_safety - 50) / 50,
            'vol_regime': vol_score, 'distribution_quality': dist_score,
        }

    @staticmethod
    def score_batch(reports, idx_df=None):
        if not reports: return {}
        factor_data = {}
        for rp in reports:
            if 'error' in rp: continue
            factor_data[rp['symbol']] = AdaptiveScorer.extract_factors(rp)
        if len(factor_data) < 2:
            return AdaptiveScorer._score_absolute(reports, idx_df)
        symbols = list(factor_data.keys())
        factor_names = list(AdaptiveScorer.WEIGHTS.keys())
        factor_arrays = {fn: np.array([factor_data[s].get(fn,0) for s in symbols]) for fn in factor_names}
        scores = {}
        vni_regime = AdaptiveScorer._detect_vni_regime(idx_df)
        regime_mult = {'BULL': 1.08, 'NEUTRAL': 1.00, 'WEAK': 0.92, 'BEAR': 0.85, 'CRISIS': 0.70}

        # V4.1 FIX #6: Compute absolute scores first for blending
        abs_scores = AdaptiveScorer._score_absolute(reports, idx_df)
        anchor = CFG.ABSOLUTE_SCORE_ANCHOR  # 0.5 = 50% absolute + 50% cross-sectional

        for i, sym in enumerate(symbols):
            weighted_sum = 0.0; details = {}
            for fn in factor_names:
                raw = factor_arrays[fn][i]
                z = AdaptiveScorer._robust_zscore(raw, factor_arrays[fn])
                w = AdaptiveScorer.WEIGHTS[fn]
                weighted_sum += w * z
                details[fn] = {'raw': round(float(raw),4), 'z': round(z,3), 'w': w}
            cs_score = 50 + weighted_sum * 40
            cs_adj = cs_score * regime_mult.get(vni_regime, 1.0)

            # V4.1: Blend with absolute score to reduce universe-dependence
            abs_score = abs_scores.get(sym, {}).get('score', 50)
            blended = anchor * abs_score + (1 - anchor) * cs_adj

            rp = next((r for r in reports if r['symbol']==sym), None)
            sig = (rp.get('screener_signal','') or '') if rp else ''
            if 'MẠNH' in sig: blended += 4
            elif 'TÍCH LŨY' in sig and factor_data[sym].get('vol_regime',0)>0: blended += 3
            elif 'PHÂN PHỐI' in sig: blended -= 6
            final = int(np.clip(blended, 0, 100))
            rt = ('⭐⭐⭐⭐⭐ SWING BUY' if final>=80 else '⭐⭐⭐⭐ BUY & HOLD' if final>=65 else
                  '⭐⭐⭐ THEO DÕI' if final>=50 else '⭐⭐ YẾU' if final>=35 else '⭐ TRÁNH')
            fc = (rp.get('fcast',{}) if rp else {})
            scores[sym] = {'score': final, 'rating': rt, 'vni_regime': vni_regime,
                           'timing': fc.get('timing',''), 'hold_plan': fc.get('hold_plan_label',''),
                           'scoring_method': 'cross_sectional_zscore', 'factor_details': details}
        return scores

    @staticmethod
    def _score_absolute(reports, idx_df=None):
        scores = {}; vni_regime = AdaptiveScorer._detect_vni_regime(idx_df)
        regime_mult = {'BULL': 1.08, 'NEUTRAL': 1.0, 'WEAK': 0.92, 'BEAR': 0.85, 'CRISIS': 0.70}
        for rp in reports:
            if 'error' in rp: continue
            sym = rp['symbol']; f = AdaptiveScorer.extract_factors(rp)
            sc = 50
            sc += np.clip(f['sharpe'] * 10, -15, 15)
            sc += np.clip(f['win_rate'] * 10, -8, 8)
            sc += np.clip(f['forecast_ret'] * 2, -10, 10)
            sc += np.clip(f['cmf'] * 30, -7, 7)
            sc += np.clip(f['trend_quality'] * 12, -7, 7)
            sc += np.clip(f['alpha_composite'] * 10, -10, 10)
            sc += np.clip(f['hmm_regime'] * 5, -5, 5)
            sc += np.clip(f['lock_safety'] * 5, -5, 5)
            sc = sc * regime_mult.get(vni_regime, 1.0)
            sig = (rp.get('screener_signal','') or '')
            if 'MẠNH' in sig: sc += 4
            elif 'PHÂN PHỐI' in sig: sc -= 6
            final = int(np.clip(sc, 0, 100))
            rt = ('⭐⭐⭐⭐⭐ SWING BUY' if final>=80 else '⭐⭐⭐⭐ BUY & HOLD' if final>=65 else
                  '⭐⭐⭐ THEO DÕI' if final>=50 else '⭐⭐ YẾU' if final>=35 else '⭐ TRÁNH')
            fc = rp.get('fcast',{})
            scores[sym] = {'score': final, 'rating': rt, 'vni_regime': vni_regime,
                           'timing': fc.get('timing',''), 'hold_plan': fc.get('hold_plan_label',''),
                           'scoring_method': 'absolute_fallback'}
        return scores

    @staticmethod
    def _detect_vni_regime(idx_df):
        if idx_df is None or 'close' not in idx_df.columns or len(idx_df)<50: return 'NEUTRAL'
        try:
            c = idx_df['close']
            ma20 = c.rolling(20).mean().iloc[-1]; ma50 = c.rolling(50).mean().iloc[-1]
            last = c.iloc[-1]
            ret_20d = (last / c.iloc[-20] - 1) if len(c)>=20 else 0
            ret_5d = (last / c.iloc[-5] - 1) if len(c)>=5 else 0
            ret_1d = (last / c.iloc[-2] - 1) if len(c)>=2 else 0

            # V4.1 FIX #4: CRISIS = extreme short-term damage
            # Condition 1: 1-day crash AND 5-day decline
            if ret_1d < CFG.VNI_CRASH_1D_PCT and ret_5d < CFG.VNI_CRASH_5D_PCT:
                return 'CRISIS'
            # Condition 2: Severe 5-day crash alone
            if ret_5d < CFG.VNI_CRASH_5D_PCT * 1.5:  # -7.5% in 5 days
                return 'CRISIS'
            # Condition 3: Single-day capitulation (e.g. -3% flash crash)
            if ret_1d < CFG.VNI_CRASH_1D_PCT * 1.5:  # -3% in 1 day
                return 'CRISIS'
            # Existing logic
            if last > ma20 > ma50 and ret_20d > 0.02: return 'BULL'
            elif last < ma20 < ma50 and ret_20d < -0.02: return 'BEAR'
            elif last < ma20 and ret_5d < -0.01: return 'WEAK'
            else: return 'NEUTRAL'
        except: return 'NEUTRAL'

# ==============================================================
# ORCHESTRATOR — V4
# ==============================================================

class QuantPipeline:
    def __init__(self, cfg=None):
        self.cfg = cfg or CFG
        log.info("QuantPipeline V4 initialized (9 modules + AdaptiveScorer)")

    def analyze(self, sym, df, sig=None, scores=None, idx_df=None, vni_regime='NEUTRAL'):
        if df is None or len(df)<30: return {'symbol':sym, 'error':'Insufficient data'}
        p = df['close']
        r = {'symbol':sym, 'date':datetime.now().strftime('%Y-%m-%d %H:%M'),
             'n':len(df), 'range':f"{str(df.index.min())[:10]} -> {str(df.index.max())[:10]}",
             'screener_signal':sig, 'screener_scores':scores,
             '_price_dates': df.index.tolist(), '_price_close': p.tolist()}
        log.info(f"  [{sym}] M1: Distribution..."); r['dist'] = DistributionAnalyzer.full_test(p)
        log.info(f"  [{sym}] M2: Statistics..."); r['stats'] = StatEngine.returns(p); r['vol'] = StatEngine.vol_regime(p); r['ac'] = StatEngine.autocorr(p)
        log.info(f"  [{sym}] M3: ARIMA..."); r['arima'] = ARIMAEngine.fit(p)
        log.info(f"  [{sym}] M4: GARCH..."); r['garch'] = GARCHEngine.fit(p)
        # V4.1: Pass symbol for stable HMM seeding
        log.info(f"  [{sym}] M5: HMM (multivariate)..."); r['hmm'] = HMMEngine.fit(df, symbol=sym)
        log.info(f"  [{sym}] M6: Alpha..."); r['alpha'] = AlphaEngine.extract(df, idx_df)
        log.info(f"  [{sym}] M7: Structure (swing SR)..."); r['sr'] = StructureEngine.sr_levels(df); r['trend'] = StructureEngine.trend(df); r['flow'] = StructureEngine.flow(df)
        # V4.1: Pass GARCH result + symbol + VNI regime to MC
        log.info(f"  [{sym}] M8: Forecast (GARCH-MC)..."); r['fcast'] = FcastEngine.ensemble(p, garch_result=r.get('garch'), symbol=sym, vni_regime=vni_regime)
        log.info(f"  [{sym}] M9: Risk...")
        sl = RiskEng.stops(df); r['sl'] = sl
        fc_hold = r.get('fcast',{}).get('hold_plan_sessions', CFG.SWING_DEFAULT)
        if sl and sl.get('sl_swing'):
            r['pos'] = RiskEng.sizing(sl['entry'], sl['sl_swing'], hold_sessions=fc_hold)
        rs = r.get('stats',{})
        if rs.get('win_rate_pct',0)>0:
            r['kelly'] = RiskEng.kelly(rs['win_rate_pct']/100, rs.get('avg_win_pct',0)/100, rs.get('avg_loss_pct',1)/100)
        # Scoring deferred to batch() for cross-sectional z-scores
        r['rec'] = {'score': 50, 'rating': '⭐⭐⭐ PENDING', 'vni_regime': 'NEUTRAL'}
        return r

    def batch(self, data, scr_df=None, idx_df=None):
        # V4.1: Detect VNI regime EARLY so each stock's forecast respects it
        vni_regime = AdaptiveScorer._detect_vni_regime(idx_df)
        log.info(f"\n  [VNI REGIME] {vni_regime}")
        if vni_regime == 'CRISIS':
            log.warning("  ⚠️ CRISIS DETECTED — circuit breaker active, all timing restricted")

        reports = []
        for i,(sym,df) in enumerate(data.items()):
            log.info(f"\n[{i+1}/{len(data)}] === {sym} ===")
            sig, scores = None, None
            if scr_df is not None and not scr_df.empty:
                mc = [c for c in scr_df.columns if 'ma' in c.lower() or 'symbol' in c.lower()]
                if mc:
                    m = scr_df[scr_df[mc[0]]==sym]
                    if not m.empty:
                        row = m.iloc[0]
                        sc2 = [c for c in scr_df.columns if 'tin hieu' in c.lower() or 'adj' in c.lower()]
                        sig = str(row[sc2[0]]) if sc2 else None
            reports.append(self.analyze(sym, df, sig, scores, idx_df, vni_regime=vni_regime))
        # V4: Batch scoring
        log.info("\n  [BATCH] Adaptive cross-sectional scoring...")
        batch_scores = AdaptiveScorer.score_batch(reports, idx_df)
        for rp in reports:
            if 'error' in rp: continue
            if rp['symbol'] in batch_scores:
                rp['rec'] = batch_scores[rp['symbol']]
        reports.sort(key=lambda x:x.get('rec',{}).get('score',0), reverse=True)
        return reports

    # ==============================================================
    # COMMENTARY ENGINE — V4.1: Thesis-driven narrative
    # ==============================================================
    # Structure per stock:
    #   ① Sức khỏe tài chính — thesis + metrics as PROOF + explain WHY
    #   ② Trạng thái kỹ thuật — vol regime, trend, money flow as STORY
    #   ③ Mô hình thống kê — HMM/Forecast/Kelly as CONVICTION measure
    #   ④ Rủi ro — GARCH, distribution, lock risk as HONEST WARNING
    #   Kết luận — action + levels + sizing in VND
    # ==============================================================

    def generate_commentary(self, r):
        if 'error' in r:
            return 'Không đủ dữ liệu để phân tích.'

        sym = r.get('symbol', '?')
        rs = r.get('stats', {}); vl = r.get('vol', {}); hm = r.get('hmm', {})
        fc = r.get('fcast', {}); ga = r.get('garch', {}).get('garch', {})
        eg = r.get('garch', {}).get('egarch', {})
        di = r.get('dist', {}).get('verdict', {}); ft = r.get('dist', {}).get('fat_tail', {})
        sl = r.get('sl', {}); al = r.get('alpha', {}); ky = r.get('kelly', {})
        fl = r.get('flow', {}); tr = r.get('trend', {}); rc = r.get('rec', {})
        sig = r.get('screener_signal', '') or ''
        lock = fc.get('lock_risk', {}); pos = r.get('pos', {})

        score = rc.get('score', 0); rating = rc.get('rating', '')
        parts = []

        # ── HEADER ──
        parts.append(f'{sym} — Score {score} ({rating})')

        # ── ① SỨC KHỎE TÀI CHÍNH ──
        parts.append(self._section_health(rs, ky))

        # ── ② TRẠNG THÁI KỸ THUẬT ──
        parts.append(self._section_technical(vl, tr, fl, al))

        # ── ③ MÔ HÌNH & ĐỘ TIN CẬY ──
        parts.append(self._section_models(hm, fc, ky, sig))

        # ── ④ RỦI RO ──
        parts.append(self._section_risk(ga, eg, di, ft, rs, lock, sl))

        # ── KẾT LUẬN ──
        parts.append(self._section_conclusion(sym, rc, fc, sl, pos, lock))

        return '\n'.join(parts)

    # -----------------------------------------------------------------
    def _section_health(self, rs, ky):
        """① Thesis: stock financial health — prove with Sharpe/Sortino/Calmar."""
        ar = rs.get('ann_return_pct', 0)
        sh = rs.get('sharpe', 0)
        so = rs.get('sortino', 0)
        cal = rs.get('calmar', 0)
        mdd = rs.get('max_dd_pct', 0)
        wr = rs.get('win_rate_pct', 0)
        pf = rs.get('profit_factor', 0)

        lines = []

        # Determine thesis
        if sh >= 1.5:
            thesis = 'Hiệu suất tài chính xuất sắc'
        elif sh >= 0.5:
            thesis = 'Hiệu suất tài chính tích cực'
        elif sh >= 0:
            thesis = 'Hiệu suất tài chính trung bình'
        else:
            thesis = 'Hiệu suất tài chính yếu'

        lines.append(f'① {thesis}')

        # Build narrative
        narr = f'AnnRet {ar:.1f}%'
        if sh >= 0.5:
            narr += f' với Sharpe {sh:.2f} — mỗi đơn vị rủi ro chấp nhận, cổ phiếu trả lại {sh:.1f} đơn vị lợi nhuận'
            if so > sh * 1.2:
                narr += f'. Sortino {so:.2f} cao hơn Sharpe đáng kể — phần lớn biến động là biến động tăng, không phải giảm'
        elif sh >= 0:
            narr += f', Sharpe {sh:.2f} — lợi nhuận có bù được rủi ro nhưng chưa ấn tượng'
        else:
            narr += f', Sharpe {sh:.2f} — lợi nhuận KHÔNG bù được rủi ro đã chấp nhận. '
            narr += f'Nói thẳng: gửi tiết kiệm có thể tốt hơn nếu chỉ nhìn lịch sử'

        # Calmar — relate to MaxDD
        if cal > 3:
            narr += f'. Calmar {cal:.1f} — lợi nhuận năm gấp {cal:.1f}x mức thua xấu nhất (MaxDD {mdd:.1f}%)'
        elif mdd < -15:
            narr += f'. MaxDD {mdd:.1f}% — từng mất hơn 15% từ đỉnh, đây là mức drawdown nặng'
        elif mdd < -10:
            narr += f'. MaxDD {mdd:.1f}% — cần chuẩn bị tâm lý cho drawdown 2 chữ số'

        # Win rate + profit factor context
        if wr > 0 and pf > 0:
            if pf > 1.5 and wr > 55:
                narr += f'. WinRate {wr:.0f}%, Profit Factor {pf:.2f} — thắng nhiều hơn thua, và khi thắng thì lớn hơn khi thua'
            elif pf < 1:
                narr += f'. Profit Factor {pf:.2f} (<1) — khi thua, mức lỗ trung bình lớn hơn mức lãi. Đây là dấu hiệu cần cải thiện'

        lines.append(narr + '.')
        return '\n'.join(lines)

    # -----------------------------------------------------------------
    def _section_technical(self, vl, tr, fl, al):
        """② Thesis: current technical state — vol, trend, flow, RS."""
        lines = []

        # Determine thesis from vol regime
        vol_rg = vl.get('regime', '')
        vol_ratio = vl.get('vol_ratio', 1)
        vol_pct = vl.get('vol_pctile', 50)

        if vol_rg == 'CONTRACTION':
            thesis = 'Đang "nén lò xo" — biến động thấp bất thường'
            vol_detail = (f'Vol Ratio {vol_ratio:.3f} tại Percentile {vol_pct:.1f}% — '
                          f'biến động 10 ngày chỉ bằng {vol_ratio:.0%} so với 60 ngày, '
                          f'thấp hơn {100-vol_pct:.0f}% lịch sử. '
                          f'Im lặng kéo dài thường đi trước bùng nổ — đây là giai đoạn tích lũy')
        elif vol_rg == 'EXPANSION':
            thesis = 'Biến động đang mở rộng — thị trường "nóng"'
            vol_detail = (f'Vol Ratio {vol_ratio:.3f} — biến động ngắn hạn cao hơn {vol_ratio:.0%} so với trung hạn. '
                          f'Cần cẩn trọng: biến động cao = cả cơ hội lẫn rủi ro lớn hơn')
        else:
            thesis = 'Biến động bình thường'
            vol_detail = f'Vol Ratio {vol_ratio:.3f} — không có tín hiệu bất thường về biến động'

        lines.append(f'② {thesis}')
        narr = vol_detail

        # Trend quality
        tr_label = tr.get('label', '')
        er = tr.get('er', 0)
        slope = tr.get('slope_pct', 0)
        if er > 0.5:
            narr += f'. Xu hướng rõ ràng (ER {er:.2f}, slope {slope:+.1f}%) — giá di chuyển có hướng, không lắc lư ngẫu nhiên'
        elif er > 0.3:
            narr += f'. Xu hướng nhẹ (ER {er:.2f}) — có hướng nhưng nhiễu khá nhiều'
        else:
            narr += f'. Chưa có xu hướng rõ (ER {er:.2f}) — giá lắc lư không phương hướng'

        # Money flow
        cmf = fl.get('cmf', 0)
        if cmf > 0.1:
            narr += f'. Dòng tiền tích cực (CMF {cmf:+.3f}) — tiền thông minh đang tích lũy'
        elif cmf < -0.1:
            narr += f'. Dòng tiền tiêu cực (CMF {cmf:+.3f}) — có dấu hiệu phân phối, tiền đang rút ra'
        elif abs(cmf) <= 0.1 and cmf != 0:
            narr += f'. Dòng tiền trung tính (CMF {cmf:+.3f})'

        # Relative Strength vs VNINDEX
        cs = al.get('cross_sectional', {})
        if cs:
            rs20 = cs.get('rs_20d_pct', 0)
            beta = cs.get('beta', 1)
            alpha_ann = cs.get('alpha_ann_pct', 0)
            if rs20 > 3:
                narr += f'. RS 20d +{rs20:.1f}% vs VNINDEX — vượt trội thị trường rõ rệt, dòng tiền ưu tiên mã này'
            elif rs20 > 0:
                narr += f'. RS 20d +{rs20:.1f}% vs VNINDEX — mạnh hơn thị trường nhẹ'
            elif rs20 < -3:
                narr += f'. RS 20d {rs20:.1f}% — yếu hơn VNINDEX đáng kể, dòng tiền đang né tránh'
            if abs(alpha_ann) > 5:
                narr += f' (alpha {alpha_ann:+.1f}%/năm, beta {beta:.2f})'

        # Momentum/RSI context
        mom = al.get('momentum', {})
        rsi = mom.get('rsi', 50)
        if rsi > 70:
            narr += f'. RSI {rsi:.0f} — vùng quá mua, cẩn trọng đuổi giá'
        elif rsi < 30:
            narr += f'. RSI {rsi:.0f} — vùng quá bán, có thể hồi kỹ thuật'

        lines.append(narr + '.')
        return '\n'.join(lines)

    # -----------------------------------------------------------------
    def _section_models(self, hm, fc, ky, sig=''):
        """③ Thesis: statistical models agree/disagree — HMM, forecast, Kelly."""
        lines = []

        hmm_cur = hm.get('current', '')
        hmm_prob = hm.get('prob_pct', 0)
        hmm_warning = hm.get('warning', '')
        ens_ret = fc.get('ensemble_ret_pct', 0)
        conf = fc.get('confidence', 0)
        mc_src = fc.get('mc', {}).get('vol_source', '')
        prob_up = fc.get('mc', {}).get('prob_up', 50)

        # Thesis from agreement level
        agree_count = 0
        if 'BULL' in hmm_cur and hmm_prob >= 60: agree_count += 1
        if ens_ret > 1: agree_count += 1
        if ky.get('has_edge'): agree_count += 1
        if conf >= 66: agree_count += 1

        if agree_count >= 3:
            thesis = 'Mô hình thống kê đồng thuận TĂNG'
        elif agree_count >= 2:
            thesis = 'Tín hiệu đang hình thành nhưng chưa chín muồi'
        elif agree_count == 1:
            thesis = 'Tín hiệu yếu — mô hình chưa đồng thuận'
        else:
            thesis = 'Không có tín hiệu rõ từ mô hình'

        lines.append(f'③ {thesis}')

        # HMM narrative
        narr = ''
        hmm_n = hm.get('n_features', 1)
        if 'BULL' in hmm_cur:
            narr = f'HMM {hmm_cur} xác suất {hmm_prob:.0f}%'
            if hmm_prob >= 80:
                narr += ' — xác suất áp đảo, regime ổn định'
            elif hmm_prob >= 60:
                narr += ' — nghiêng tăng nhưng chưa áp đảo'
            else:
                narr += ' — chưa chắc chắn, regime có thể đổi'
            if hmm_n > 1:
                narr += f' (phân tích {hmm_n} chiều: giá, vol, volume, momentum)'
        elif 'BEAR' in hmm_cur:
            narr = f'HMM {hmm_cur} xác suất {hmm_prob:.0f}% — cổ phiếu đang trong trạng thái giảm'
        else:
            narr = f'HMM {hmm_cur} — thị trường đi ngang, chưa rõ hướng'

        # Forecast
        narr += f'. Forecast {ens_ret:+.2f}% trong {fc.get("horizon",10)} phiên'
        if conf >= 66:
            narr += f', Confidence {conf:.0f}% — các mô hình khá đồng thuận'
        elif conf >= 50:
            narr += f', Confidence {conf:.0f}% — tín hiệu đang hình thành, chưa chín muồi'
        else:
            narr += f', Confidence {conf:.0f}% — ngang tung đồng xu, mô hình chưa rõ hướng'

        if mc_src:
            narr += f' (MC dùng {mc_src} vol)'

        # Kelly
        if ky.get('has_edge'):
            k_full = ky.get('full_pct', 0)
            k_half = ky.get('half_pct', 0)
            edge = ky.get('edge_pct', 0)
            narr += (f'. Kelly xác nhận edge thực sự: Half Kelly {k_half:.1f}%, '
                     f'edge {edge:.1f}% — kỳ vọng dương trên mỗi giao dịch')
        elif ky:
            narr += '. Kelly âm — chưa có lợi thế thống kê, xác suất thắng chưa đủ bù thua'

        # Screener confluence
        if sig and sig != 'nan':
            if 'MẠNH' in sig:
                narr += f'. Screener cũng xác nhận tín hiệu MẠNH (đa tiêu chí kỹ thuật) — tăng niềm tin'
            elif 'TÍCH LŨY' in sig:
                narr += f'. Screener: đang TÍCH LŨY — phù hợp với giai đoạn chờ breakout'
            elif 'PHÂN PHỐI' in sig:
                narr += f'. ⚠️ Screener cảnh báo PHÂN PHỐI — mâu thuẫn với tín hiệu tăng, cần thận trọng'

        lines.append(narr + '.')
        return '\n'.join(lines)

    # -----------------------------------------------------------------
    def _section_risk(self, ga, eg, di, ft, rs, lock, sl):
        """④ Honest risk warnings with explanation of WHY each metric matters."""
        lines = []
        items = []

        # GARCH
        persist = ga.get('persistence', 0)
        alpha_g = ga.get('alpha', 0)
        shock = ga.get('shock_sensitivity', '')
        if persist > 0.9:
            p_label = 'IGARCH-like' if persist > 0.97 else 'cao'
            item = f'GARCH persistence {persist:.3f} ({p_label})'
            if alpha_g > 0.15:
                item += f', Alpha {alpha_g:.4f} — nếu có tin xấu bất ngờ, biến động cao sẽ kéo dài rất lâu'
            else:
                item += ' — cú sốc biến động tắt chậm'
            hl = ga.get('half_life')
            if hl and hl > 0:
                item += f' (half-life {hl:.0f} phiên)'
            items.append(item)

        # EGARCH leverage
        if eg and not eg.get('error'):
            gamma = eg.get('gamma', 0)
            if gamma < -0.05:
                items.append(f'EGARCH gamma {gamma:.3f} — tin xấu tăng biến động MẠNH hơn tin tốt (leverage effect)')

        # Distribution
        if not di.get('is_gaussian', True):
            reject_n = di.get('reject_count', 0)
            kurt = ft.get('excess_kurtosis', 0)
            sev = ft.get('severity', '')
            item = f'Phân phối NON-GAUSSIAN ({reject_n}/5 test bác bỏ)'
            if kurt > 5:
                item += f', kurtosis {kurt:.1f} ({sev}) — đuôi rất dày, "thiên nga đen" xảy ra thường hơn mô hình dự đoán'
            elif kurt > 1:
                item += f', kurtosis {kurt:.1f} — cần dùng VaR non-parametric thay vì Gaussian'
            items.append(item)

        # VaR / CVaR context
        var95 = rs.get('VaR_95', 0)
        cvar95 = rs.get('CVaR_95', 0)
        if var95 and cvar95:
            items.append(f'VaR 95%: {var95:.2f}% (95/100 ngày lỗ không quá mức này), CVaR: {cvar95:.2f}% (trung bình 5 ngày tệ nhất)')

        # Lock risk
        prob_l3 = lock.get('prob_loss_gt_3pct', 0)
        prob_lock = lock.get('prob_loss_in_lock_pct', 0)
        mae = sl.get('mae_lock_10pct', 0)
        if prob_l3 > 15 or prob_lock > 50:
            item = f'T+3 lock: {prob_lock:.0f}% khả năng lỗ khi chưa bán được'
            if prob_l3 > 20:
                item += f', {prob_l3:.0f}% khả năng lỗ >3%'
            if mae:
                item += f'. Lịch sử: 90% trường hợp DD trong lock < {abs(mae):.1f}%'
            items.append(item)

        # Build section
        if items:
            lines.append('④ Rủi ro cần lưu ý')
            for item in items:
                lines.append(item + '.')
        else:
            lines.append('④ Rủi ro: Không có cảnh báo đặc biệt.')

        return '\n'.join(lines)

    # -----------------------------------------------------------------
    def _section_conclusion(self, sym, rc, fc, sl, pos, lock):
        """Conclusion: verdict + concrete levels + sizing."""
        score = rc.get('score', 0)
        rating = rc.get('rating', '')
        ens_ret = fc.get('ensemble_ret_pct', 0)
        hold_plan = fc.get('hold_plan_label', '')

        entry = sl.get('entry', 0)
        sl_swing = sl.get('sl_swing', 0)
        tp2 = sl.get('tp_2r', 0)
        tp3 = sl.get('tp_3r', 0)
        shares = pos.get('shares', 0)
        value = pos.get('value', 0)

        lines = []

        # Verdict paragraph
        risk_per = entry - sl_swing if entry and sl_swing else 0
        rr_2 = (tp2 - entry) / risk_per if risk_per > 0 and tp2 else 0
        rr_3 = (tp3 - entry) / risk_per if risk_per > 0 and tp3 else 0
        risk_pct = risk_per / entry * 100 if entry > 0 else 0

        if score >= 80:
            verdict = (f'Tổng kết {sym}: Đủ điều kiện SWING BUY — tín hiệu mạnh, rủi ro kiểm soát được')
        elif score >= 65:
            verdict = (f'Tổng kết {sym}: Có thể MUA với quản lý chặt — tín hiệu tích cực nhưng cần kỷ luật')
        elif score >= 50:
            verdict = (f'Tổng kết {sym}: THEO DÕI — tín hiệu đang hình thành nhưng chưa đủ mạnh để vào lệnh')
        elif score >= 35:
            verdict = (f'Tổng kết {sym}: TÍN HIỆU YẾU — không nên mở vị thế mới')
        else:
            verdict = (f'Tổng kết {sym}: TRÁNH — rủi ro cao hơn cơ hội')

        lines.append(verdict + '.')

        # Strategy line
        if entry and sl_swing and tp2 and score >= 50:
            strat = f'Chiến lược: '
            if score >= 65:
                strat += f'Entry {entry:,.0f}'
            else:
                strat += f'Alert tại {entry:,.0f}, vào khi có xác nhận'
            strat += f'. SL {sl_swing:,.0f} (-{risk_pct:.0f}%) — TP {tp2:,.0f}/{tp3:,.0f} (R:R 1:{rr_2:.1f}/{rr_3:.1f})'

            if shares > 0 and value > 0:
                strat += f'. Size: {shares:,.0f} cp ({value/1e6:,.1f}M VND)'

            if hold_plan:
                strat += f'. Nắm giữ: {hold_plan}'

            lines.append(strat + '.')

        # Priority ranking hint
        if score >= 75:
            lines.append(f'Ưu tiên CAO trong danh sách.')
        elif score >= 60:
            lines.append(f'Ưu tiên TRUNG BÌNH — cần thêm xác nhận.')

        return '\n'.join(lines)

    def summary_table(self, reports):
        rows = []
        for r in reports:
            if 'error' in r: continue
            rc = r.get('rec',{}); rs = r.get('stats',{}); fc = r.get('fcast',{})
            sl = r.get('sl',{}); ps = r.get('pos',{}); vl = r.get('vol',{})
            hm = r.get('hmm',{}); lock = fc.get('lock_risk',{}); unlock = fc.get('unlock',{})
            rows.append({
                'Symbol':r['symbol'], 'Score':rc.get('score',0), 'Rating':rc.get('rating',''),
                'Timing':rc.get('timing',''), 'HoldPlan':rc.get('hold_plan',''),
                'VNI':rc.get('vni_regime',''), 'Screener':r.get('screener_signal',''),
                'Sharpe':rs.get('sharpe',0), 'WinRate':rs.get('win_rate_pct',0),
                'MaxDD':rs.get('max_dd_pct',0), 'AnnRet':rs.get('ann_return_pct',0),
                'HMM':hm.get('current',''), 'VolReg':vl.get('regime',''),
                'Forecast':fc.get('consensus',''), 'EnsRet%':fc.get('ensemble_ret_pct',0),
                'Conf':fc.get('confidence',0),
                'MCVol':fc.get('mc',{}).get('vol_source',''),
                'LockDD%':lock.get('max_dd_lock_pct',0), 'PLoss3%':lock.get('prob_loss_gt_3pct',0),
                'Entry':sl.get('entry',0), 'SL':sl.get('sl_swing',0),
                'TP2R':sl.get('tp_2r',0), 'TP3R':sl.get('tp_3r',0),
                'Shares':ps.get('shares',0), 'Value':ps.get('value',0),
                'Method':rc.get('scoring_method',''),
                'Analysis': self.generate_commentary(r),
            })
        return pd.DataFrame(rows)

# ==============================================================
# FORECAST LOGGER
# ==============================================================

class ForecastLogger:
    COLUMNS = ['symbol','run_date','horizon','forecast_dir','forecast_ret_pct',
               'forecast_price','consensus','confidence','actual_price','actual_ret_pct','hit']
    def __init__(self, filepath='forecast_log.csv'):
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            pd.DataFrame(columns=self.COLUMNS).to_csv(self.filepath, index=False)
    def log_reports(self, reports, horizon=None):
        if horizon is None: horizon = CFG.SWING_DEFAULT
        rows = []; now = datetime.now().strftime('%Y-%m-%d %H:%M')
        for rp in reports:
            if 'error' in rp: continue
            fc = rp.get('fcast',{}); 
            if not fc: continue
            rows.append({
                'symbol': rp['symbol'], 'run_date': now, 'horizon': horizon,
                'forecast_dir': 'UP' if 'TANG' in fc.get('consensus','') else ('DOWN' if 'GIAM' in fc.get('consensus','') else 'FLAT'),
                'forecast_ret_pct': fc.get('ensemble_ret_pct',0), 'forecast_price': fc.get('ensemble_price',0),
                'consensus': fc.get('consensus',''), 'confidence': fc.get('confidence',0),
                'actual_price': '', 'actual_ret_pct': '', 'hit': '',
            })
        if rows:
            pd.DataFrame(rows).to_csv(self.filepath, mode='a', header=False, index=False)
            log.info(f"Logged {len(rows)} forecasts")
    def evaluate(self, bridge, days_back=30):
        try: df = pd.read_csv(self.filepath)
        except: return {}
        if df.empty: return {}
        df['run_date'] = pd.to_datetime(df['run_date'])
        pending = df[(df['actual_price']=='')|(df['actual_price'].isna())]
        pending = pending[pending['run_date']<=(datetime.now()-timedelta(days=5))]
        if pending.empty: return {}
        updated = 0
        for sym in pending['symbol'].unique():
            try:
                sd = bridge.fetch_ohlcv([sym], days=60, delay=0.1)
                if sym not in sd: continue
                prices = sd[sym]['close']
                mask = (df['symbol']==sym)&((df['actual_price']=='')|(df['actual_price'].isna()))
                for idx in df[mask].index:
                    run_dt = df.loc[idx,'run_date']; horizon = int(df.loc[idx,'horizon'])
                    target_dt = run_dt + timedelta(days=horizon+2)
                    fp = prices[prices.index>=target_dt.strftime('%Y-%m-%d')]
                    if fp.empty: continue
                    actual = float(fp.iloc[0])
                    fcast_price = float(df.loc[idx,'forecast_price'])
                    entry_price = fcast_price / (1+float(df.loc[idx,'forecast_ret_pct'])/100)
                    actual_ret_pct = (actual/entry_price - 1)*100
                    fcast_dir = df.loc[idx,'forecast_dir']
                    hit = 'Y' if ((fcast_dir=='UP' and actual_ret_pct>0) or
                                  (fcast_dir=='DOWN' and actual_ret_pct<0) or
                                  (fcast_dir=='FLAT' and abs(actual_ret_pct)<1)) else 'N'
                    df.loc[idx,'actual_price'] = round(actual,0)
                    df.loc[idx,'actual_ret_pct'] = round(actual_ret_pct,2)
                    df.loc[idx,'hit'] = hit; updated += 1
            except Exception as e: log.warning(f"Eval {sym}: {e}")
        df.to_csv(self.filepath, index=False)
        evaluated = df[df['hit'].isin(['Y','N'])]
        if evaluated.empty: return {'total': 0}
        total = len(evaluated); hits = len(evaluated[evaluated['hit']=='Y'])
        return {'total': total, 'hits': hits, 'hit_rate_pct': round(hits/total*100,1)}

# ==============================================================
# FORECAST SHEET EXPORT — for report.py chart rendering
# ==============================================================

def _export_forecast_sheet(reports, output_path):
    """Append a 'Forecast' sheet to the existing quant_report.xlsx.
    
    Builds per-symbol rows with:
      - Historical close prices (last 30 trading days)
      - MC-predicted median/CI for the forecast horizon
      - Predicted = ARIMA fitted values (backfill) + MC median (forward)
    
    This sheet is consumed by report.py → Charts.price_forecast() and Charts.ml_forecast().
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment

    rows = []
    for rp in reports:
        if 'error' in rp:
            continue
        sym = rp['symbol']
        fc = rp.get('fcast', {})
        mc_path = fc.get('mc_path')
        if not mc_path:
            continue

        dates_raw = rp.get('_price_dates', [])
        closes_raw = rp.get('_price_close', [])
        if not dates_raw or not closes_raw:
            continue

        # Take last N_HIST trading days for context
        N_HIST = 30
        dates_hist = dates_raw[-N_HIST:]
        closes_hist = closes_raw[-N_HIST:]

        # ARIMA fitted values as "predicted" for historical zone
        arima = rp.get('arima', {})
        fitted_raw = arima.get('fitted', [])
        if fitted_raw and len(fitted_raw) >= N_HIST:
            pred_hist = list(fitted_raw[-N_HIST:])
        else:
            # Fallback: use close as predicted (perfect fit) for historical
            pred_hist = list(closes_hist)

        # Forecast zone: MC median path
        median_path = mc_path['median']
        ci_upper = mc_path['upper']
        ci_lower = mc_path['lower']
        horizon = len(median_path)

        # Generate forecast dates (next business days)
        last_date = pd.Timestamp(dates_hist[-1])
        fc_dates = pd.bdate_range(start=last_date + pd.Timedelta(days=1), periods=horizon)

        forecast_start_idx = len(dates_hist)

        # Historical rows
        for i, (d, c, p) in enumerate(zip(dates_hist, closes_hist, pred_hist)):
            rows.append({
                'Symbol': sym,
                'Date': pd.Timestamp(d),
                'Close': round(float(c), 2),
                'Predicted': round(float(p), 2),
                'Upper': None,
                'Lower': None,
                'Forecast_Start': forecast_start_idx if i == 0 else None,
            })

        # Forecast rows (close = NaN → tells report.py this is OOS)
        for i in range(horizon):
            rows.append({
                'Symbol': sym,
                'Date': fc_dates[i],
                'Close': None,  # NaN signals forecast zone
                'Predicted': round(float(median_path[i]), 2),
                'Upper': round(float(ci_upper[i]), 2),
                'Lower': round(float(ci_lower[i]), 2),
                'Forecast_Start': None,
            })

    if not rows:
        log.info("No forecast data to export")
        return

    df_fc = pd.DataFrame(rows)

    # Write to existing workbook
    try:
        wb = load_workbook(output_path)
        # Remove old Forecast sheet if exists
        if 'Forecast' in wb.sheetnames:
            del wb['Forecast']
        ws = wb.create_sheet('Forecast')
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Forecast'

    # Header
    HEADER_FONT = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    from openpyxl.styles import PatternFill
    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    headers = ['Symbol', 'Date', 'Close', 'Predicted', 'Upper', 'Lower', 'Forecast_Start']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for ri, row in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=row['Symbol'])
        ws.cell(row=ri, column=2, value=row['Date'])
        ws.cell(row=ri, column=3, value=row['Close'])
        ws.cell(row=ri, column=4, value=row['Predicted'])
        ws.cell(row=ri, column=5, value=row['Upper'])
        ws.cell(row=ri, column=6, value=row['Lower'])
        ws.cell(row=ri, column=7, value=row['Forecast_Start'])

    # Format date column
    for ri in range(2, len(rows) + 2):
        ws.cell(row=ri, column=2).number_format = 'YYYY-MM-DD'

    # Column widths
    for ci, w in enumerate([10, 14, 12, 12, 12, 12, 14], 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A2'
    wb.save(output_path)
    log.info(f"  Forecast sheet exported: {len(rows)} rows for {df_fc['Symbol'].nunique()} symbols → {output_path}")


# ==============================================================
# ENHANCED EXCEL FALLBACK (if backtest_engine not available)
# ==============================================================

def _export_enhanced_inline(summary_df, output_path):
    """Fallback enhanced Excel export with bold/red formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active; ws.title = 'Summary'

    FONT_BOLD = Font(name='Arial', bold=True, size=10, color='000000')
    FONT_RED_BOLD = Font(name='Arial', bold=True, size=10, color='FF0000')
    FONT_NORMAL = Font(name='Arial', size=10, color='000000')
    FONT_HEADER = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    FILL_HEADER = PatternFill('solid', fgColor='2E4057')
    FILL_ALT = PatternFill('solid', fgColor='F5F7FA')
    FILL_GREEN = PatternFill('solid', fgColor='D4EDDA')
    FILL_RED = PatternFill('solid', fgColor='F8D7DA')
    FILL_YELLOW = PatternFill('solid', fgColor='FFF3CD')
    BORDER_THIN = Border(bottom=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'))

    cols = list(summary_df.columns)
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.border = BORDER_THIN

    for ri, (_, row) in enumerate(summary_df.iterrows(), 2):
        score_val = row.get('Score', 0)
        row_fill = FILL_GREEN if score_val >= 80 else (FILL_RED if score_val < 35 else (FILL_YELLOW if score_val < 50 else (FILL_ALT if ri % 2 == 0 else None)))
        for ci, col_name in enumerate(cols, 1):
            val = row.get(col_name, '')
            cell = ws.cell(row=ri, column=ci)
            cell.border = BORDER_THIN
            if col_name == 'Analysis':
                cell.value = str(val) if val else ''
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                has_danger = any(kw in str(val) for kw in ['KHÔNG VÀO', 'TRÁNH', '🔴'])
                cell.font = Font(name='Arial', size=9, color='CC0000') if has_danger else Font(name='Arial', bold=True, size=9, color='1A1A1A') if any(kw in str(val) for kw in ['Entry ', 'Chiến lược:']) else Font(name='Arial', size=9, color='333333')
            else:
                cell.value = val
                cell.alignment = Alignment(horizontal='center', vertical='top')
                if col_name in ('Score', 'Rating', 'Timing', 'HoldPlan'): cell.font = FONT_BOLD
                elif col_name in ('Entry', 'SL', 'TP2R', 'TP3R'): cell.font = FONT_RED_BOLD
                else: cell.font = FONT_NORMAL
            if row_fill: cell.fill = row_fill

    col_widths = {'Symbol': 9, 'Score': 8, 'Rating': 22, 'Timing': 35, 'HoldPlan': 16, 'Analysis': 110}
    for ci, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col_name, 12)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(summary_df)+1}"
    wb.save(output_path)
    log.info(f"\n📁 Enhanced Excel exported: {output_path}")

# ==============================================================
# MAIN
# ==============================================================

# ==============================================================
# BUY-ONLY SUMMARY EXCEL — Tổng hợp chỉ mã khuyến nghị MUA
# ==============================================================

def _export_buy_summary(summary_df, output_path='buy_summary.xlsx', vni_regime='NEUTRAL'):
    """Xuất Excel tổng hợp chỉ các mã MUA (Score >= 50 & Forecast = TANG)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import numpy as np

    buy_mask = (summary_df['Score'] >= 50) & (summary_df['Forecast'].str.contains('TANG', na=False))
    buy_df = summary_df[buy_mask].copy().sort_values('Score', ascending=False).reset_index(drop=True)
    total_all = len(summary_df); total_buy = len(buy_df)
    now = datetime.now(); date_str = now.strftime('%d/%m/%Y'); time_str = now.strftime('%H:%M')

    if total_buy == 0:
        log.info(f"Không có mã nào đủ điều kiện MUA — skip buy summary export.")
        return output_path

    FONT_TITLE = Font(name='Arial', bold=True, size=14, color='1F3864')
    FONT_SUB = Font(name='Arial', size=10, color='666666')
    FONT_SECTION = Font(name='Arial', bold=True, size=12, color='1F3864')
    FONT_HEADER = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    FONT_BOLD = Font(name='Arial', bold=True, size=10, color='000000')
    FONT_NORMAL = Font(name='Arial', size=10, color='333333')
    FONT_RED = Font(name='Arial', bold=True, size=10, color='CC0000')
    FONT_GREEN = Font(name='Arial', bold=True, size=10, color='006100')
    FONT_KPI_LABEL = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    FONT_KPI_VAL = Font(name='Arial', bold=True, size=16, color='FFFFFF')
    FILL_DARK = PatternFill('solid', fgColor='1F3864')
    FILL_GREEN_KPI = PatternFill('solid', fgColor='375623')
    FILL_ORANGE_KPI = PatternFill('solid', fgColor='C55A11')
    FILL_TEAL_KPI = PatternFill('solid', fgColor='2E75B6')
    FILL_HEADER = PatternFill('solid', fgColor='2E4057')
    FILL_ALT1 = PatternFill('solid', fgColor='F5F7FA')
    FILL_ALT2 = PatternFill('solid', fgColor='FFFFFF')
    FILL_BDS = PatternFill('solid', fgColor='FCE4D6')
    FILL_BANK = PatternFill('solid', fgColor='DDEBF7')
    FILL_OTHER = PatternFill('solid', fgColor='F2F2F2')
    FILL_LIGHT_GREEN = PatternFill('solid', fgColor='E2EFDA')
    FILL_LIGHT_RED = PatternFill('solid', fgColor='FFC7CE')
    FILL_YELLOW = PatternFill('solid', fgColor='FFF3CD')
    BORDER = Border(bottom=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                    left=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'))
    ALIGN_C = Alignment(horizontal='center', vertical='center')
    ALIGN_L = Alignment(horizontal='left', vertical='center')

    wb = Workbook()

    # ── SHEET 1: TỔNG QUAN MUA ──
    ws1 = wb.active; ws1.title = 'Tổng Quan Mua'
    ws1.merge_cells('B2:K2')
    ws1['B2'].value = f'TỔNG HỢP KHUYẾN NGHỊ MUA — BOT QUANT {date_str}'; ws1['B2'].font = FONT_TITLE
    ws1.merge_cells('B3:K3')
    ws1['B3'].value = f'VNI: {vni_regime}  |  Tạo ngày {date_str} {time_str}  |  Lọc: Score ≥ 50 & Forecast = TĂNG'; ws1['B3'].font = FONT_SUB

    # KPI boxes
    kpi_data = [
        ('B', 'C', 'Tổng mã scan', total_all, FILL_DARK),
        ('D', 'E', 'Mã MUA', total_buy, FILL_GREEN_KPI),
        ('F', 'G', 'Tỉ lệ MUA', f'{total_buy}/{total_all} ({total_buy/total_all*100:.0f}%)', FILL_ORANGE_KPI),
        ('H', 'I', 'Avg Score (MUA)', f'{buy_df["Score"].mean():.1f}', FILL_TEAL_KPI),
        ('J', 'K', 'VNI Regime', vni_regime, FILL_ORANGE_KPI),
    ]
    for cl, cr, label, val, fill in kpi_data:
        ws1.merge_cells(f'{cl}5:{cr}5')
        c = ws1[f'{cl}5']; c.value = label; c.font = FONT_KPI_LABEL; c.fill = fill; c.alignment = ALIGN_C; ws1[f'{cr}5'].fill = fill
        ws1.merge_cells(f'{cl}6:{cr}6')
        c = ws1[f'{cl}6']; c.value = val; c.font = FONT_KPI_VAL; c.fill = fill; c.alignment = ALIGN_C; ws1[f'{cr}6'].fill = fill

    # Rating breakdown
    ws1['B8'].value = 'Phân Bổ Theo Rating'; ws1['B8'].font = FONT_SECTION
    for ci, h in enumerate(['Rating', 'Số mã', 'Tỉ lệ', 'Avg Score', 'Mã'], 2):
        c = ws1.cell(row=9, column=ci, value=h); c.font = FONT_HEADER; c.fill = FILL_DARK; c.alignment = ALIGN_C; c.border = BORDER
    row_idx = 10
    for rt in ['⭐⭐⭐⭐⭐ SWING BUY', '⭐⭐⭐⭐ BUY & HOLD', '⭐⭐⭐ THEO DÕI']:
        rt_df = buy_df[buy_df['Rating'] == rt]
        if rt_df.empty: continue
        cnt = len(rt_df); fill_r = FILL_LIGHT_GREEN if '⭐⭐⭐⭐' in rt else FILL_ALT1
        for ci, val in enumerate([rt, cnt, f'{cnt/total_buy*100:.0f}%', f'{rt_df["Score"].mean():.1f}', ', '.join(rt_df['Symbol'].tolist())], 2):
            c = ws1.cell(row=row_idx, column=ci, value=val); c.font = FONT_BOLD if ci <= 4 else FONT_NORMAL; c.fill = fill_r; c.border = BORDER
            c.alignment = ALIGN_C if ci <= 4 else ALIGN_L
        row_idx += 1

    # Timing breakdown
    row_idx += 1
    ws1.cell(row=row_idx, column=2).value = 'Phân Bổ Timing (mã MUA)'; ws1.cell(row=row_idx, column=2).font = FONT_SECTION
    row_idx += 1
    for ci, h in enumerate(['Timing', 'Số mã', 'Tỉ lệ', 'Mã'], 2):
        c = ws1.cell(row=row_idx, column=ci, value=h); c.font = FONT_HEADER; c.fill = FILL_DARK; c.alignment = ALIGN_C; c.border = BORDER
    row_idx += 1
    for prefix, fill_t in [('🟢 VÀO NGAY', FILL_LIGHT_GREEN), ('🟡 CHỜ PULLBACK', FILL_YELLOW), ('🟡 THEO DÕI', FILL_YELLOW), ('🔴 KHÔNG VÀO', FILL_LIGHT_RED)]:
        t_df = buy_df[buy_df['Timing'].str.contains(prefix.split(' —')[0], na=False)]
        if t_df.empty: continue
        cnt = len(t_df)
        for ci, val in enumerate([prefix, cnt, f'{cnt/total_buy*100:.0f}%', ', '.join(t_df['Symbol'].tolist())], 2):
            c = ws1.cell(row=row_idx, column=ci, value=val); c.font = FONT_NORMAL; c.fill = fill_t; c.border = BORDER; c.alignment = ALIGN_C if ci <= 4 else ALIGN_L
        row_idx += 1

    for col, w in {'A':2,'B':22,'C':12,'D':18,'E':12,'F':55}.items():
        ws1.column_dimensions[col].width = w
    for col in 'GHIJK':
        ws1.column_dimensions[col].width = 14

    # ── SHEET 2: DANH SÁCH MÃ MUA ──
    ws2 = wb.create_sheet('Danh Sách Mua')
    ws2.merge_cells('B2:S2')
    ws2['B2'].value = f'{total_buy} MÃ KHUYẾN NGHỊ MUA — BOT QUANT {date_str}'; ws2['B2'].font = FONT_TITLE
    ws2.merge_cells('B3:S3')
    ws2['B3'].value = 'Lọc: Score ≥ 50 & Forecast TĂNG  |  Sắp xếp theo Score giảm dần'; ws2['B3'].font = FONT_SUB

    detail_cols = [
        ('Mã','Symbol',9), ('Score','Score',8), ('Rating','Rating',22), ('HMM','HMM',12),
        ('Forecast','Forecast',12), ('EnsRet%','EnsRet%',10), ('Conf%','Conf',8),
        ('Entry','Entry',10), ('SL','SL',10), ('TP1(2R)','TP2R',10), ('TP2(3R)','TP3R',10),
        ('→SL%','_sl_pct',8), ('→TP1%','_tp1_pct',8), ('R:R','_rr',8),
        ('Timing','Timing',35), ('Nắm giữ','HoldPlan',14), ('Shares','Shares',10), ('Value(M)','_value_m',12),
    ]
    ws2.column_dimensions['A'].width = 2
    for ci, (header, _, width) in enumerate(detail_cols, 2):
        c = ws2.cell(row=4, column=ci, value=header); c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = ALIGN_C; c.border = BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = width

    for ri, (_, row) in enumerate(buy_df.iterrows(), 5):
        entry = row.get('Entry',0) or 0; sl_p = row.get('SL',0) or 0; tp2r = row.get('TP2R',0) or 0
        sl_pct = (sl_p - entry)/entry if entry > 0 else 0
        tp1_pct = (tp2r - entry)/entry if entry > 0 else 0
        risk = entry - sl_p; rr = (tp2r - entry)/risk if risk > 0 else 0
        score = row.get('Score',0)
        fill_row = FILL_LIGHT_GREEN if score >= 65 else (FILL_ALT1 if ri % 2 == 0 else FILL_ALT2)
        for ci, (_, col_key, _) in enumerate(detail_cols, 2):
            c = ws2.cell(row=ri, column=ci); c.border = BORDER; c.fill = fill_row; c.alignment = ALIGN_C
            if col_key == '_sl_pct': c.value = sl_pct; c.number_format = '0.0%'; c.font = FONT_RED
            elif col_key == '_tp1_pct': c.value = tp1_pct; c.number_format = '0.0%'; c.font = FONT_GREEN
            elif col_key == '_rr': c.value = round(rr,2) if rr > 0 else ''; c.font = FONT_BOLD
            elif col_key == '_value_m': c.value = round((row.get('Value',0) or 0)/1e6,1); c.font = FONT_NORMAL
            elif col_key == 'EnsRet%': v = row.get(col_key,0); c.value = v; c.number_format = '0.00%'; c.font = FONT_GREEN if v > 0 else FONT_RED
            elif col_key == 'Conf': v = row.get(col_key,0); c.value = v/100 if v > 1 else v; c.number_format = '0%'; c.font = FONT_NORMAL
            elif col_key in ('Entry','SL','TP2R','TP3R'): c.value = row.get(col_key,''); c.font = FONT_RED if col_key in ('Entry','SL') else FONT_GREEN; c.number_format = '#,##0'
            elif col_key == 'Score': c.value = score; c.font = FONT_BOLD
            elif col_key == 'Shares': c.value = row.get(col_key,0); c.number_format = '#,##0'; c.font = FONT_NORMAL
            else:
                c.value = row.get(col_key,'')
                if col_key == 'Rating': c.font = FONT_BOLD; c.alignment = ALIGN_L
                elif col_key == 'Timing': c.font = FONT_RED if '🔴' in str(row.get(col_key,'')) else (FONT_GREEN if '🟢' in str(row.get(col_key,'')) else FONT_NORMAL); c.alignment = ALIGN_L
                else: c.font = FONT_NORMAL

    ws2.freeze_panes = 'B5'
    ws2.auto_filter.ref = f"B4:{get_column_letter(len(detail_cols)+1)}{4+total_buy}"
    wb.save(output_path)
    log.info(f"\n📁 Buy Summary exported: {output_path} ({total_buy}/{total_all} mã MUA)")
    return output_path


def _export_recommendations_parquet(summary_df, output_path='cache/recommendations_latest.parquet',
                                     vni_regime='NEUTRAL'):
    """
    Xuất recommendations ra parquet để valuation.py đọc.

    OUTPUT SCHEMA (chuẩn cho valuation.py):
        symbol     : Mã CK (uppercase)
        action     : BUY / HOLD / SELL
        score      : 0-100 (giữ nguyên scale từ pipeline)
        entry      : Giá vào lệnh (VND)
        stop       : Giá cắt lỗ (VND)
        target     : Giá chốt lời (VND, dùng TP2R)
        regime     : HMM regime + VNI regime (concatenated)
        rating     : Rating gốc từ pipeline (⭐⭐⭐⭐⭐ SWING BUY ...)
        forecast   : TANG / GIAM / SIDEWAYS
        confidence : 0-1 từ forecast engine

    LOGIC ACTION (consistent với _export_buy_summary):
        - BUY  : Score >= 50 AND Forecast contains 'TANG'
        - SELL : Score < 35 OR Forecast contains 'GIAM'
        - HOLD : còn lại

    Args:
        summary_df: output của QuantPipeline.summary_table()
        output_path: đường dẫn parquet (default: cache/recommendations_latest.parquet)
        vni_regime: VNI regime hiện tại (đính kèm metadata)
    """
    if summary_df is None or summary_df.empty:
        log.warning("Summary df trống — skip recommendations parquet export.")
        return None

    # Map score → action
    def _score_to_action(row):
        score = row.get('Score', 0)
        forecast = str(row.get('Forecast', '')).upper()
        # BUY: score cao + forecast tăng
        if score >= 50 and 'TANG' in forecast:
            return 'BUY'
        # SELL: score quá thấp hoặc forecast giảm
        if score < 35 or 'GIAM' in forecast:
            return 'SELL'
        return 'HOLD'

    # Build standardized dataframe
    out = pd.DataFrame()
    out['symbol'] = summary_df['Symbol'].astype(str).str.upper().str.strip()
    out['action'] = summary_df.apply(_score_to_action, axis=1)
    out['score'] = pd.to_numeric(summary_df['Score'], errors='coerce').fillna(0)
    out['entry'] = pd.to_numeric(summary_df['Entry'], errors='coerce')
    out['stop'] = pd.to_numeric(summary_df.get('SL', 0), errors='coerce')
    # Dùng TP2R làm target chính (conservative); TP3R có thể dùng làm extended target
    out['target'] = pd.to_numeric(summary_df.get('TP2R', 0), errors='coerce')
    out['target_extended'] = pd.to_numeric(summary_df.get('TP3R', 0), errors='coerce')

    # Regime: combine HMM + VNI để valuation/report có context đầy đủ
    hmm_regime = summary_df.get('HMM', pd.Series(['']*len(summary_df))).astype(str)
    vni_col = summary_df.get('VNI', pd.Series([vni_regime]*len(summary_df))).astype(str)
    out['regime'] = hmm_regime + ' | VNI:' + vni_col

    # Metadata cho audit & valuation context
    out['rating'] = summary_df.get('Rating', '').astype(str)
    out['forecast'] = summary_df.get('Forecast', '').astype(str)
    out['confidence'] = pd.to_numeric(summary_df.get('Conf', 0), errors='coerce').fillna(0)
    out['sharpe'] = pd.to_numeric(summary_df.get('Sharpe', 0), errors='coerce').fillna(0)
    out['vol_regime'] = summary_df.get('VolReg', '').astype(str)
    out['screener_signal'] = summary_df.get('Screener', '').astype(str)

    # Timestamp khi pipeline chạy (để valuation.py biết freshness)
    out['generated_at'] = datetime.now().isoformat(timespec='seconds')
    out['vni_regime'] = vni_regime

    # Cleanup: drop rows không có symbol hoặc entry
    before = len(out)
    out = out.dropna(subset=['symbol', 'entry'])
    out = out[out['symbol'].str.len() > 0]
    after = len(out)
    if before != after:
        log.info(f"  Dropped {before-after} rows missing symbol/entry")

    # Ensure output directory exists
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save
    try:
        out.to_parquet(output_path, index=False, engine='pyarrow')
    except ImportError:
        # Fallback: dùng fastparquet nếu pyarrow không có
        try:
            out.to_parquet(output_path, index=False, engine='fastparquet')
        except ImportError:
            log.error("Cần install pyarrow hoặc fastparquet: pip install pyarrow")
            return None

    # Log summary
    action_counts = out['action'].value_counts().to_dict()
    log.info(f"✓ Recommendations parquet → {output_path}")
    log.info(f"  Total: {len(out)} symbols | "
             f"BUY: {action_counts.get('BUY', 0)} | "
             f"HOLD: {action_counts.get('HOLD', 0)} | "
             f"SELL: {action_counts.get('SELL', 0)}")

    return output_path


def run_pipeline(excel_path=None, symbols=None, account_size=500_000_000,
                 lookback_days=252, output_excel='quant_report.xlsx',
                 buy_summary_excel='buy_summary.xlsx', source=None):
    # source=None → dùng fallback chain KBS → MSN (khuyến nghị)
    # source='KBS' hoặc 'MSN' → ép dùng 1 source duy nhất
    bridge = ScreenerBridge(source=source); scr_df = pd.DataFrame()
    if excel_path:
        syms, scr_df = bridge.read_screener_excel(excel_path)
        if not symbols: symbols = syms
    elif not symbols:
        syms, scr_df = bridge.read_screener_excel(); symbols = syms
    if not symbols: log.error("Khong co symbols!"); return [], pd.DataFrame()
    log.info(f"\n{'='*60}\n  QUANT PIPELINE V4 — {len(symbols)} stocks\n  Account: {account_size:,.0f} VND\n{'='*60}")
    data = bridge.fetch_ohlcv(symbols, days=lookback_days)
    idx_df = bridge.fetch_index('VNINDEX', days=lookback_days)
    cfg = QuantConfig(); cfg.ACCOUNT_SIZE = account_size
    pipe = QuantPipeline(cfg)
    reports = pipe.batch(data, scr_df, idx_df)
    summary = pipe.summary_table(reports)
    if not summary.empty:
        print(f"\n{'='*80}\n  SUMMARY — Top {len(summary)} by Score\n{'='*80}")
        print(summary.to_string(index=False))
    if output_excel and not summary.empty:
        try:
            from backtest_engine import export_enhanced_excel
            export_enhanced_excel(summary, output_path=output_excel)
        except ImportError:
            # Fallback: inline enhanced export
            try:
                _export_enhanced_inline(summary, output_excel)
            except Exception as e: log.warning(f"Export error: {e}")
        except Exception as e: log.warning(f"Export error: {e}")
    # ── Export Forecast sheet (for report.py chart rendering) ──
    if output_excel and reports:
        try:
            _export_forecast_sheet(reports, output_excel)
        except Exception as e: log.warning(f"Forecast sheet export error: {e}")
    # ── Export Buy-Only Summary ──
    if buy_summary_excel and not summary.empty:
        try:
            vni_reg = AdaptiveScorer._detect_vni_regime(idx_df)
            _export_buy_summary(summary, buy_summary_excel, vni_regime=vni_reg)
        except Exception as e: log.warning(f"Buy summary export error: {e}")
    # ── Export Recommendations Parquet (cho valuation.py) ──
    if not summary.empty:
        try:
            vni_reg = AdaptiveScorer._detect_vni_regime(idx_df)
            _export_recommendations_parquet(summary,
                                            output_path='cache/recommendations_latest.parquet',
                                            vni_regime=vni_reg)
        except Exception as e: log.warning(f"Recommendations parquet export error: {e}")
    try:
        flogger = ForecastLogger('forecast_log.csv')
        flogger.log_reports(reports); flogger.evaluate(bridge)
    except Exception as e: log.warning(f"Forecast logger: {e}")
    return reports, summary

if __name__ == '__main__':
    print("""
    ╔══════════════════════════════════════════════╗
    ║    QUANT PIPELINE V4.1 — HOSE Swing Trading  ║
    ║    FIX: HMM stable seed, GARCH-MC vol cap,   ║
    ║         VNI crisis breaker, blended scoring,  ║
    ║         forecast cap, regime-aware timing     ║
    ╚══════════════════════════════════════════════╝
    """)
    if '--symbols' in sys.argv:
        idx = sys.argv.index('--symbols')
        s = sys.argv[idx+1].split(',') if idx+1<len(sys.argv) else []
        run_pipeline(symbols=[x.strip().upper() for x in s])
    else:
        run_pipeline()
